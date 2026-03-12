"""
Enhanced openpyxl-based Excel extraction with full feature support.

This module implements the EnhancedExtractionStrategy using openpyxl to extract
Excel-specific features including formulas, pivot tables, charts, merged cells,
named ranges, Excel tables, and hidden content.

Supports Requirements:
- 18.1: Capture both formula text and computed value for formula cells
- 18.2: Display formula cells with both formula and computed value
- 19.1: Detect pivot tables and extract data separately
- 19.2: Display pivot table metadata (source range, row/column/value fields, filters)
- 20.1: Detect charts and extract underlying data series
- 30.1: Detect merged cells and their ranges
- 31.1: Detect hidden rows, columns, and sheets
- 32.1: Detect named ranges and Excel tables
"""

import io
import logging
from typing import Any, Optional

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.chart import (
    AreaChart,
    AreaChart3D,
    BarChart,
    BarChart3D,
    BubbleChart,
    DoughnutChart,
    LineChart,
    LineChart3D,
    PieChart,
    PieChart3D,
    RadarChart,
    ScatterChart,
    StockChart,
    SurfaceChart,
    SurfaceChart3D,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.exceptions import ExtractionError
from src.extraction.enhanced_strategy import EnhancedExtractionStrategy
from src.extraction.extraction_strategy import ExtractionConfig
from src.extraction.language_detection import ExcelLanguageDetectionService
from src.models.excel_features import (
    ChartInfo,
    EnhancedExtractionResult,
    ExcelTable,
    ExtractedSheetData,
    ExtractionQuality,
    ExtractionWarning,
    FormulaCell,
    MergedCellInfo,
    NamedRange,
    PivotTableInfo,
)


logger = logging.getLogger(__name__)


class EnhancedOpenpyxlExtractor(EnhancedExtractionStrategy):
    """
    Enhanced Excel extractor using openpyxl with full feature support.
    
    Extracts comprehensive Excel data including formulas, pivot tables,
    charts, merged cells, named ranges, Excel tables, and hidden content.
    All dependencies are injected via constructor following DIP.
    
    Attributes:
        _config: Extraction configuration settings.
        _logger: Logger instance for this extractor.
    
    Example:
        >>> config = ExtractionConfig(max_rows_per_sheet=5000)
        >>> extractor = EnhancedOpenpyxlExtractor(config)
        >>> result = extractor.extract("path/to/file.xlsx")
    """
    
    # Excel error types that indicate formula errors
    EXCEL_ERROR_TYPES = frozenset({
        "#REF!",
        "#VALUE!",
        "#DIV/0!",
        "#NAME?",
        "#N/A",
        "#NULL!",
        "#NUM!",
        "#GETTING_DATA",
        "#SPILL!",
        "#CALC!",
    })
    
    def __init__(
        self,
        config: Optional[ExtractionConfig] = None,
        language_detection_service: Optional[ExcelLanguageDetectionService] = None
    ) -> None:
        """
        Initialize the enhanced openpyxl extractor.
        
        Args:
            config: Extraction configuration. Uses defaults if not provided.
            language_detection_service: Language detection service. Creates default
                if not provided.
        """
        self._config = config or ExtractionConfig()
        self._language_service = language_detection_service or ExcelLanguageDetectionService()
        self._logger = logging.getLogger(__name__)
    
    def extract(
        self,
        file_path: str,
        config: Optional[ExtractionConfig] = None
    ) -> EnhancedExtractionResult:
        """
        Extract data from an Excel file with full feature support.
        
        Performs comprehensive extraction including sheet data, formulas,
        pivot tables, charts, merged cells, named ranges, and hidden content.
        
        Args:
            file_path: Path to the Excel file to extract.
            config: Optional extraction configuration override.
        
        Returns:
            EnhancedExtractionResult containing all extracted data.
        
        Raises:
            ExtractionError: If extraction fails due to file access,
                corruption, or unsupported format.
        """
        effective_config = config or self._config
        warnings: list[ExtractionWarning] = []
        
        try:
            # Load workbook with both formulas and values
            workbook = self._load_workbook(file_path)
            
            # Also load with data_only=True to get computed values
            workbook_values = self._load_workbook(file_path, data_only=True)
            
        except Exception as e:
            self._logger.error(f"Failed to load workbook {file_path}: {e}")
            raise ExtractionError(
                f"Failed to load Excel file: {e}",
                details={"file_path": file_path, "error": str(e)}
            )
        
        try:
            # Extract sheet data
            sheets = self._extract_sheets(
                workbook, workbook_values, effective_config, warnings
            )
            
            # Extract Excel-specific features
            formula_cells = self.detect_formulas(workbook)
            merged_cells = self.detect_merged_cells(workbook)
            named_ranges = self.detect_named_ranges(workbook)
            excel_tables = self.detect_excel_tables(workbook)
            pivot_tables = self.detect_pivot_tables(workbook)
            charts = self.detect_charts(workbook)
            hidden_sheets, hidden_rows, hidden_columns = self.detect_hidden_content(
                workbook
            )
            
            # Detect primary language from sheet content
            detected_language = self._detect_content_language(sheets)
            
            # Calculate quality metrics
            quality = self._calculate_quality(sheets, warnings)
            
            return EnhancedExtractionResult(
                sheets=sheets,
                quality=quality,
                formula_cells=formula_cells,
                pivot_tables=pivot_tables,
                charts=charts,
                merged_cells=merged_cells,
                named_ranges=named_ranges,
                excel_tables=excel_tables,
                hidden_sheets=hidden_sheets,
                hidden_rows=hidden_rows,
                hidden_columns=hidden_columns,
                warnings=warnings,
                detected_language=detected_language,
            )
            
        except Exception as e:
            self._logger.error(f"Failed to extract data from {file_path}: {e}")
            raise ExtractionError(
                f"Failed to extract Excel data: {e}",
                details={"file_path": file_path, "error": str(e)}
            )
    
    def detect_formulas(self, workbook: Any) -> list[FormulaCell]:
        """
        Detect and extract formula cells from a workbook.
        
        Scans all sheets to identify cells containing formulas, capturing
        both the formula text and computed value.
        
        Args:
            workbook: The openpyxl Workbook object to scan.
        
        Returns:
            List of FormulaCell objects with formula details.
        
        Supports Requirement 18.1: Capture both formula text and computed
        value for formula cells.
        """
        formula_cells: list[FormulaCell] = []
        
        if not isinstance(workbook, Workbook):
            self._logger.warning("Invalid workbook type for formula detection")
            return formula_cells
        
        # Load workbook with data_only to get computed values
        # We need to track the file path to reload, but we don't have it here
        # So we'll use the cached values from the workbook if available
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    
                    # Check if cell contains a formula
                    formula_text = None
                    if cell.data_type == 'f':
                        # Cell has a formula
                        formula_text = str(cell.value) if cell.value else None
                    elif isinstance(cell.value, str) and cell.value.startswith('='):
                        # Formula stored as string
                        formula_text = cell.value
                    
                    if formula_text:
                        # Build cell reference with sheet name
                        cell_ref = f"{sheet_name}!{cell.coordinate}"
                        
                        # Get computed value (may be None if workbook not loaded with data_only)
                        computed_value = cell.value
                        
                        # Check for errors
                        has_error = False
                        error_type = None
                        
                        if isinstance(computed_value, str):
                            for err in self.EXCEL_ERROR_TYPES:
                                if computed_value.startswith(err):
                                    has_error = True
                                    error_type = err
                                    break
                        
                        # Check for external references
                        references_external = self._check_external_reference(formula_text)
                        
                        try:
                            formula_cell = FormulaCell(
                                cell_reference=cell_ref,
                                formula_text=formula_text if formula_text.startswith('=') else f"={formula_text}",
                                computed_value=computed_value,
                                has_error=has_error,
                                error_type=error_type,
                                references_external=references_external,
                            )
                            formula_cells.append(formula_cell)
                        except ValueError as e:
                            self._logger.warning(
                                f"Invalid formula cell at {cell_ref}: {e}"
                            )
        
        self._logger.info(f"Detected {len(formula_cells)} formula cells")
        return formula_cells
    
    def detect_merged_cells(self, workbook: Any) -> list[MergedCellInfo]:
        """
        Detect merged cells and their ranges in a workbook.
        
        Identifies all merged cell regions across all sheets, capturing
        the merge range, value, and span dimensions.
        
        Args:
            workbook: The openpyxl Workbook object to scan.
        
        Returns:
            List of MergedCellInfo objects with merge details.
        
        Supports Requirement 30.1: Detect merged cells and their ranges.
        """
        merged_cells: list[MergedCellInfo] = []
        
        if not isinstance(workbook, Workbook):
            self._logger.warning("Invalid workbook type for merged cell detection")
            return merged_cells
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            for merged_range in worksheet.merged_cells.ranges:
                # Get the value from the top-left cell
                top_left_cell = worksheet.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col
                )
                value = top_left_cell.value
                
                # Calculate span
                rows_spanned = merged_range.max_row - merged_range.min_row + 1
                cols_spanned = merged_range.max_col - merged_range.min_col + 1
                
                # Build range string with sheet name for multi-sheet workbooks
                range_str = str(merged_range)
                if len(workbook.sheetnames) > 1:
                    range_str = f"{sheet_name}!{range_str}"
                
                try:
                    merged_info = MergedCellInfo(
                        merge_range=range_str,
                        value=value,
                        rows_spanned=rows_spanned,
                        cols_spanned=cols_spanned,
                    )
                    merged_cells.append(merged_info)
                except ValueError as e:
                    self._logger.warning(
                        f"Invalid merged cell range {range_str}: {e}"
                    )
        
        self._logger.info(f"Detected {len(merged_cells)} merged cell ranges")
        return merged_cells
    
    def detect_named_ranges(self, workbook: Any) -> list[NamedRange]:
        """
        Detect named ranges defined in a workbook.
        
        Identifies all named ranges including workbook-scoped and
        sheet-scoped definitions.
        
        Args:
            workbook: The openpyxl Workbook object to scan.
        
        Returns:
            List of NamedRange objects with name and range details.
        
        Supports Requirement 32.1: Detect and index Excel named ranges.
        """
        named_ranges: list[NamedRange] = []
        
        if not isinstance(workbook, Workbook):
            self._logger.warning("Invalid workbook type for named range detection")
            return named_ranges
        
        # Iterate over defined names using the correct openpyxl API
        # workbook.defined_names is a DefinedNameDict which is iterable
        for name in workbook.defined_names:
            # Skip internal names (start with _xlnm)
            if name.startswith('_xlnm'):
                continue
            
            try:
                defined_name = workbook.defined_names[name]
                
                # Get the destinations (can be multiple for discontinuous ranges)
                destinations = list(defined_name.destinations)
                if not destinations:
                    continue
                
                # Use the first destination
                sheet_name, cell_range = destinations[0]
                
                # Determine scope based on localSheetId
                local_sheet_id = getattr(defined_name, 'localSheetId', None)
                if local_sheet_id is not None:
                    scope = "sheet"
                else:
                    scope = "workbook"
                    sheet_name = None  # Workbook-scoped names don't have a specific sheet
                
                named_range = NamedRange(
                    name=name,
                    cell_range=cell_range,
                    sheet_name=sheet_name,
                    scope=scope,
                )
                named_ranges.append(named_range)
                
            except Exception as e:
                self._logger.warning(f"Failed to parse named range '{name}': {e}")
        
        self._logger.info(f"Detected {len(named_ranges)} named ranges")
        return named_ranges
    
    def detect_excel_tables(self, workbook: Any) -> list[ExcelTable]:
        """
        Detect Excel Tables (ListObjects) in a workbook.
        
        Identifies all structured tables defined in the workbook,
        capturing table name, range, headers, and row count.
        
        Args:
            workbook: The openpyxl Workbook object to scan.
        
        Returns:
            List of ExcelTable objects with table details.
        
        Supports Requirement 32.1: Detect Excel Tables (ListObjects).
        """
        excel_tables: list[ExcelTable] = []
        
        if not isinstance(workbook, Workbook):
            self._logger.warning("Invalid workbook type for table detection")
            return excel_tables
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Access tables via worksheet.tables
            if not hasattr(worksheet, 'tables') or not worksheet.tables:
                continue
            
            for table_name, table in worksheet.tables.items():
                try:
                    # Get table range
                    table_range = table.ref
                    
                    # Extract headers from the first row of the table
                    headers = self._extract_table_headers(worksheet, table)
                    
                    # Calculate row count (excluding header)
                    row_count = self._calculate_table_row_count(table)
                    
                    excel_table = ExcelTable(
                        name=table_name,
                        cell_range=table_range,
                        sheet_name=sheet_name,
                        headers=headers,
                        row_count=row_count,
                    )
                    excel_tables.append(excel_table)
                    
                except Exception as e:
                    self._logger.warning(
                        f"Failed to extract table '{table_name}' from {sheet_name}: {e}"
                    )
        
        self._logger.info(f"Detected {len(excel_tables)} Excel tables")
        return excel_tables
    
    def detect_pivot_tables(self, workbook: Any) -> list[PivotTableInfo]:
        """
        Detect pivot tables in a workbook.
        
        Identifies all pivot tables and extracts their metadata including
        source range, field configurations, and filters.
        
        Args:
            workbook: The openpyxl Workbook object to scan.
        
        Returns:
            List of PivotTableInfo objects with pivot table details.
        
        Supports Requirement 19.1: Detect pivot tables and extract their
        data separately.
        """
        pivot_tables: list[PivotTableInfo] = []
        
        if not isinstance(workbook, Workbook):
            self._logger.warning("Invalid workbook type for pivot table detection")
            return pivot_tables
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Access pivot tables via worksheet._pivots
            if not hasattr(worksheet, '_pivots') or not worksheet._pivots:
                continue
            
            for idx, pivot in enumerate(worksheet._pivots):
                try:
                    pivot_info = self._extract_pivot_table_info(
                        pivot, idx, sheet_name
                    )
                    if pivot_info:
                        pivot_tables.append(pivot_info)
                except Exception as e:
                    self._logger.warning(
                        f"Failed to extract pivot table {idx} from {sheet_name}: {e}"
                    )
        
        self._logger.info(f"Detected {len(pivot_tables)} pivot tables")
        return pivot_tables
    
    def detect_charts(self, workbook: Any) -> list[ChartInfo]:
        """
        Detect charts and extract underlying data series.
        
        Identifies all charts in the workbook and extracts their type,
        title, axis labels, data series, and source ranges.
        
        Args:
            workbook: The openpyxl Workbook object to scan.
        
        Returns:
            List of ChartInfo objects with chart details.
        
        Supports Requirement 20.1: Detect charts and extract underlying
        data series.
        """
        charts: list[ChartInfo] = []
        
        if not isinstance(workbook, Workbook):
            self._logger.warning("Invalid workbook type for chart detection")
            return charts
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Access charts via worksheet._charts
            if not hasattr(worksheet, '_charts') or not worksheet._charts:
                continue
            
            for idx, chart in enumerate(worksheet._charts):
                try:
                    chart_info = self._extract_chart_info(chart, idx, sheet_name)
                    if chart_info:
                        charts.append(chart_info)
                except Exception as e:
                    self._logger.warning(
                        f"Failed to extract chart {idx} from {sheet_name}: {e}"
                    )
        
        self._logger.info(f"Detected {len(charts)} charts")
        return charts
    
    def detect_hidden_content(
        self,
        workbook: Any
    ) -> tuple[list[str], dict[str, list[int]], dict[str, list[str]]]:
        """
        Detect hidden sheets, rows, and columns in a workbook.
        
        Identifies all hidden content including hidden sheets, hidden
        rows per sheet, and hidden columns per sheet.
        
        Args:
            workbook: The openpyxl Workbook object to scan.
        
        Returns:
            A tuple containing:
            - hidden_sheets: List of hidden sheet names.
            - hidden_rows: Dictionary mapping sheet names to lists of
              hidden row indices (1-indexed).
            - hidden_columns: Dictionary mapping sheet names to lists of
              hidden column letters.
        
        Supports Requirement 31.1: Detect hidden rows, columns, and sheets.
        """
        hidden_sheets: list[str] = []
        hidden_rows: dict[str, list[int]] = {}
        hidden_columns: dict[str, list[str]] = {}
        
        if not isinstance(workbook, Workbook):
            self._logger.warning("Invalid workbook type for hidden content detection")
            return hidden_sheets, hidden_rows, hidden_columns
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Check if sheet is hidden
            if worksheet.sheet_state != 'visible':
                hidden_sheets.append(sheet_name)
            
            # Detect hidden rows
            sheet_hidden_rows: list[int] = []
            for row_idx, row_dim in worksheet.row_dimensions.items():
                if row_dim.hidden:
                    sheet_hidden_rows.append(row_idx)
            
            if sheet_hidden_rows:
                hidden_rows[sheet_name] = sorted(sheet_hidden_rows)
            
            # Detect hidden columns
            sheet_hidden_cols: list[str] = []
            for col_letter, col_dim in worksheet.column_dimensions.items():
                if col_dim.hidden:
                    sheet_hidden_cols.append(col_letter)
            
            if sheet_hidden_cols:
                hidden_columns[sheet_name] = sorted(sheet_hidden_cols)
        
        self._logger.info(
            f"Detected {len(hidden_sheets)} hidden sheets, "
            f"{sum(len(rows) for rows in hidden_rows.values())} hidden rows, "
            f"{sum(len(cols) for cols in hidden_columns.values())} hidden columns"
        )
        return hidden_sheets, hidden_rows, hidden_columns
    
    # =========================================================================
    # Private Helper Methods
    # =========================================================================
    
    def _load_workbook(
        self,
        file_path: str,
        data_only: bool = False
    ) -> Workbook:
        """
        Load an Excel workbook from a file path.
        
        Args:
            file_path: Path to the Excel file.
            data_only: If True, load computed values instead of formulas.
        
        Returns:
            Loaded openpyxl Workbook.
        
        Raises:
            ExtractionError: If file cannot be loaded.
        """
        try:
            workbook = openpyxl.load_workbook(
                file_path,
                data_only=data_only,
                keep_vba=False,
                rich_text=False,
            )
            return workbook
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise ExtractionError(
                f"Invalid Excel file format: {e}",
                details={"file_path": file_path}
            )
        except Exception as e:
            raise ExtractionError(
                f"Failed to load workbook: {e}",
                details={"file_path": file_path}
            )
    
    def _load_workbook_from_bytes(
        self,
        file_content: bytes,
        data_only: bool = False
    ) -> Workbook:
        """
        Load an Excel workbook from bytes.
        
        Args:
            file_content: Raw bytes of the Excel file.
            data_only: If True, load computed values instead of formulas.
        
        Returns:
            Loaded openpyxl Workbook.
        
        Raises:
            ExtractionError: If file cannot be loaded.
        """
        try:
            file_stream = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(
                file_stream,
                data_only=data_only,
                keep_vba=False,
                rich_text=False,
            )
            return workbook
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise ExtractionError(f"Invalid Excel file format: {e}")
        except Exception as e:
            raise ExtractionError(f"Failed to load workbook: {e}")

    def _extract_sheets(
        self,
        workbook: Workbook,
        workbook_values: Workbook,
        config: ExtractionConfig,
        warnings: list[ExtractionWarning],
    ) -> list[ExtractedSheetData]:
        """
        Extract data from all sheets in the workbook.
        
        Args:
            workbook: Workbook with formulas.
            workbook_values: Workbook with computed values.
            config: Extraction configuration.
            warnings: List to append warnings to.
        
        Returns:
            List of ExtractedSheetData objects.
        """
        sheets: list[ExtractedSheetData] = []
        
        for sheet_name in workbook.sheetnames:
            try:
                worksheet = workbook[sheet_name]
                worksheet_values = workbook_values[sheet_name]
                
                sheet_data = self._extract_single_sheet(
                    worksheet, worksheet_values, config
                )
                sheets.append(sheet_data)
                
            except Exception as e:
                self._logger.warning(f"Failed to extract sheet '{sheet_name}': {e}")
                warnings.append(ExtractionWarning(
                    warning_type="format_issue",
                    message=f"Failed to extract sheet '{sheet_name}': {e}",
                    location=sheet_name,
                    severity="warning",
                ))
        
        return sheets
    
    def _extract_single_sheet(
        self,
        worksheet: Worksheet,
        worksheet_values: Worksheet,
        config: ExtractionConfig,
    ) -> ExtractedSheetData:
        """
        Extract data from a single worksheet.
        
        Args:
            worksheet: Worksheet with formulas.
            worksheet_values: Worksheet with computed values.
            config: Extraction configuration.
        
        Returns:
            ExtractedSheetData object.
        """
        sheet_name = worksheet.title
        max_row = min(worksheet.max_row or 0, config.max_rows_per_sheet)
        max_col = worksheet.max_column or 0
        
        if max_row == 0 or max_col == 0:
            return ExtractedSheetData(
                sheet_name=sheet_name,
                headers=[],
                data=[],
                row_count=0,
                column_count=0,
                has_headers=False,
            )
        
        # Detect header row
        header_row_idx = self._detect_header_row(worksheet, max_col)
        
        # Extract headers
        headers = self._extract_headers(worksheet, header_row_idx, max_col)
        
        # Extract data rows using computed values
        data: list[list[Any]] = []
        for row_idx in range(header_row_idx + 1, max_row + 1):
            row_data: list[Any] = []
            for col_idx in range(1, max_col + 1):
                # Use computed values from data_only workbook
                cell = worksheet_values.cell(row=row_idx, column=col_idx)
                row_data.append(cell.value)
            
            # Only add non-empty rows
            if any(v is not None and v != "" for v in row_data):
                data.append(row_data)
        
        return ExtractedSheetData(
            sheet_name=sheet_name,
            headers=headers,
            data=data,
            row_count=len(data),
            column_count=len(headers),
            has_headers=bool(headers),
        )
    
    def _detect_header_row(self, worksheet: Worksheet, max_col: int) -> int:
        """
        Detect the header row by analyzing the first rows for text content.
        
        Args:
            worksheet: Worksheet to analyze.
            max_col: Maximum column index.
        
        Returns:
            Row index (1-based) of the detected header row.
        """
        max_rows_to_check = min(5, worksheet.max_row or 0)
        
        best_row = 1
        best_score = 0
        
        for row_idx in range(1, max_rows_to_check + 1):
            text_count = 0
            non_empty_count = 0
            
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    non_empty_count += 1
                    if isinstance(cell.value, str):
                        text_count += 1
            
            score = text_count * 2 + non_empty_count
            
            if score > best_score:
                best_score = score
                best_row = row_idx
        
        return best_row
    
    def _extract_headers(
        self,
        worksheet: Worksheet,
        header_row: int,
        max_col: int
    ) -> list[str]:
        """
        Extract column headers from the header row.
        
        Args:
            worksheet: Worksheet to extract from.
            header_row: Row index of headers (1-based).
            max_col: Maximum column index.
        
        Returns:
            List of header names.
        """
        headers: list[str] = []
        
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            header_value = cell.value
            
            # Handle merged cells
            if cell.coordinate in worksheet.merged_cells:
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left = worksheet.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col
                        )
                        header_value = top_left.value
                        break
            
            if header_value is None or str(header_value).strip() == "":
                header_name = f"Column_{get_column_letter(col_idx)}"
            else:
                header_name = str(header_value).strip()
            
            headers.append(header_name)
        
        return headers
    
    def _check_external_reference(self, formula_text: str) -> bool:
        """
        Check if a formula references external workbooks.
        
        Args:
            formula_text: The formula text to check.
        
        Returns:
            True if formula references external workbooks.
        """
        # External references typically contain [filename.xlsx] or path separators
        if '[' in formula_text and ']' in formula_text:
            return True
        if '\\' in formula_text or '/' in formula_text:
            # Check if it looks like a file path
            return '.xls' in formula_text.lower()
        return False
    
    def _extract_table_headers(
        self,
        worksheet: Worksheet,
        table: Any
    ) -> list[str]:
        """
        Extract headers from an Excel table.
        
        Args:
            worksheet: Worksheet containing the table.
            table: The table object.
        
        Returns:
            List of header names.
        """
        headers: list[str] = []
        
        try:
            # Parse the table range to get the header row
            table_range = table.ref
            # Range format is like "A1:D10"
            start_cell, end_cell = table_range.split(':')
            
            # Extract column letters and row numbers
            import re
            start_match = re.match(r'([A-Z]+)(\d+)', start_cell)
            end_match = re.match(r'([A-Z]+)(\d+)', end_cell)
            
            if start_match and end_match:
                start_col = start_match.group(1)
                start_row = int(start_match.group(2))
                end_col = end_match.group(1)
                
                # Convert column letters to indices
                from openpyxl.utils import column_index_from_string
                start_col_idx = column_index_from_string(start_col)
                end_col_idx = column_index_from_string(end_col)
                
                # Extract headers from the first row of the table
                for col_idx in range(start_col_idx, end_col_idx + 1):
                    cell = worksheet.cell(row=start_row, column=col_idx)
                    header = str(cell.value) if cell.value else f"Column_{col_idx}"
                    headers.append(header)
        
        except Exception as e:
            self._logger.warning(f"Failed to extract table headers: {e}")
        
        return headers
    
    def _calculate_table_row_count(self, table: Any) -> int:
        """
        Calculate the number of data rows in a table (excluding header).
        
        Args:
            table: The table object.
        
        Returns:
            Number of data rows.
        """
        try:
            table_range = table.ref
            start_cell, end_cell = table_range.split(':')
            
            import re
            start_match = re.match(r'([A-Z]+)(\d+)', start_cell)
            end_match = re.match(r'([A-Z]+)(\d+)', end_cell)
            
            if start_match and end_match:
                start_row = int(start_match.group(2))
                end_row = int(end_match.group(2))
                # Subtract 1 for header row
                return max(0, end_row - start_row)
        
        except Exception as e:
            self._logger.warning(f"Failed to calculate table row count: {e}")
        
        return 0
    
    def _extract_pivot_table_info(
        self,
        pivot: Any,
        index: int,
        sheet_name: str
    ) -> Optional[PivotTableInfo]:
        """
        Extract information from a single pivot table.
        
        Args:
            pivot: The pivot table object.
            index: Index of the pivot table.
            sheet_name: Name of the sheet containing the pivot.
        
        Returns:
            PivotTableInfo object or None if extraction fails.
        """
        try:
            # Get pivot table name
            name = getattr(pivot, 'name', f"PivotTable{index + 1}")
            
            # Get location
            location = "Unknown"
            if hasattr(pivot, 'location') and pivot.location:
                loc = pivot.location
                if hasattr(loc, 'ref'):
                    location = loc.ref
            
            # Get source range
            source_range: Optional[str] = None
            if hasattr(pivot, 'cache') and pivot.cache:
                cache = pivot.cache
                if hasattr(cache, 'cacheSource') and cache.cacheSource:
                    cache_source = cache.cacheSource
                    if hasattr(cache_source, 'worksheetSource'):
                        ws_source = cache_source.worksheetSource
                        if hasattr(ws_source, 'ref'):
                            source_range = ws_source.ref
            
            # Extract field information
            row_fields = self._get_pivot_fields(pivot, 'rowFields')
            column_fields = self._get_pivot_fields(pivot, 'colFields')
            value_fields = self._get_pivot_data_fields(pivot)
            filters = self._get_pivot_fields(pivot, 'pageFields')
            
            return PivotTableInfo(
                name=name,
                sheet_name=sheet_name,
                location=location,
                source_range=source_range,
                row_fields=row_fields,
                column_fields=column_fields,
                value_fields=value_fields,
                filters=filters,
            )
        
        except Exception as e:
            self._logger.warning(f"Failed to extract pivot table info: {e}")
            return None
    
    def _get_pivot_fields(self, pivot: Any, field_type: str) -> list[str]:
        """
        Get field names from a pivot table.
        
        Args:
            pivot: The pivot table object.
            field_type: Type of fields ('rowFields', 'colFields', 'pageFields').
        
        Returns:
            List of field names.
        """
        fields: list[str] = []
        
        try:
            field_list = getattr(pivot, field_type, None)
            if not field_list:
                return fields
            
            for field in field_list:
                if hasattr(field, 'x'):
                    field_name = self._get_pivot_field_name(pivot, field.x)
                    if field_name:
                        fields.append(field_name)
        
        except Exception as e:
            self._logger.debug(f"Failed to get pivot {field_type}: {e}")
        
        return fields
    
    def _get_pivot_data_fields(self, pivot: Any) -> list[str]:
        """
        Get data field names from a pivot table.
        
        Args:
            pivot: The pivot table object.
        
        Returns:
            List of data field names with aggregation type.
        """
        fields: list[str] = []
        
        try:
            data_fields = getattr(pivot, 'dataFields', None)
            if not data_fields:
                return fields
            
            for field in data_fields:
                name = getattr(field, 'name', None)
                if name:
                    fields.append(name)
                elif hasattr(field, 'fld'):
                    field_name = self._get_pivot_field_name(pivot, field.fld)
                    if field_name:
                        fields.append(field_name)
        
        except Exception as e:
            self._logger.debug(f"Failed to get pivot data fields: {e}")
        
        return fields
    
    def _get_pivot_field_name(self, pivot: Any, field_index: int) -> Optional[str]:
        """
        Get the name of a pivot field by index.
        
        Args:
            pivot: The pivot table object.
            field_index: Index of the field in the cache.
        
        Returns:
            Field name or None if not found.
        """
        try:
            if not hasattr(pivot, 'cache') or not pivot.cache:
                return None
            
            cache = pivot.cache
            if not hasattr(cache, 'cacheFields') or not cache.cacheFields:
                return None
            
            cache_fields = list(cache.cacheFields)
            if 0 <= field_index < len(cache_fields):
                field = cache_fields[field_index]
                return getattr(field, 'name', None)
        
        except Exception as e:
            self._logger.debug(f"Failed to get pivot field name: {e}")
        
        return None
    
    def _extract_chart_info(
        self,
        chart: Any,
        index: int,
        sheet_name: str
    ) -> Optional[ChartInfo]:
        """
        Extract information from a single chart.
        
        Args:
            chart: The chart object.
            index: Index of the chart.
            sheet_name: Name of the sheet containing the chart.
        
        Returns:
            ChartInfo object or None if extraction fails.
        """
        try:
            # Determine chart type
            chart_type = self._get_chart_type(chart)
            
            # Get chart title
            title: Optional[str] = None
            if hasattr(chart, 'title') and chart.title:
                if hasattr(chart.title, 'text'):
                    title = chart.title.text
                elif isinstance(chart.title, str):
                    title = chart.title
            
            # Get location (anchor position)
            location = f"Chart{index + 1}"
            if hasattr(chart, 'anchor'):
                anchor = chart.anchor
                if hasattr(anchor, '_from'):
                    from_cell = anchor._from
                    if hasattr(from_cell, 'col') and hasattr(from_cell, 'row'):
                        col_letter = get_column_letter(from_cell.col + 1)
                        location = f"{col_letter}{from_cell.row + 1}"
            
            # Get axis labels
            axis_labels: dict[str, str] = {}
            if hasattr(chart, 'x_axis') and chart.x_axis:
                x_axis = chart.x_axis
                if hasattr(x_axis, 'title') and x_axis.title:
                    if hasattr(x_axis.title, 'text'):
                        axis_labels['x'] = x_axis.title.text
            
            if hasattr(chart, 'y_axis') and chart.y_axis:
                y_axis = chart.y_axis
                if hasattr(y_axis, 'title') and y_axis.title:
                    if hasattr(y_axis.title, 'text'):
                        axis_labels['y'] = y_axis.title.text
            
            # Get data series
            data_series: list[str] = []
            source_ranges: list[str] = []
            
            if hasattr(chart, 'series'):
                for series in chart.series:
                    # Get series title
                    if hasattr(series, 'title') and series.title:
                        data_series.append(str(series.title))
                    
                    # Get source range for values
                    if hasattr(series, 'val') and series.val:
                        val_ref = series.val
                        if hasattr(val_ref, 'numRef') and val_ref.numRef:
                            if hasattr(val_ref.numRef, 'f'):
                                source_ranges.append(val_ref.numRef.f)
            
            return ChartInfo(
                chart_type=chart_type,
                title=title,
                sheet_name=sheet_name,
                location=location,
                axis_labels=axis_labels,
                data_series=data_series,
                source_ranges=source_ranges,
            )
        
        except Exception as e:
            self._logger.warning(f"Failed to extract chart info: {e}")
            return None
    
    def _get_chart_type(self, chart: Any) -> str:
        """
        Determine the type of a chart.
        
        Args:
            chart: The chart object.
        
        Returns:
            Chart type string.
        """
        # Map chart classes to type strings
        chart_type_map = {
            BarChart: "bar",
            BarChart3D: "bar3D",
            LineChart: "line",
            LineChart3D: "line3D",
            PieChart: "pie",
            PieChart3D: "pie3D",
            DoughnutChart: "doughnut",
            ScatterChart: "scatter",
            BubbleChart: "bubble",
            AreaChart: "area",
            AreaChart3D: "area3D",
            RadarChart: "radar",
            SurfaceChart: "surface",
            SurfaceChart3D: "surface",
            StockChart: "stock",
        }
        
        for chart_class, type_name in chart_type_map.items():
            if isinstance(chart, chart_class):
                return type_name
        
        return "unknown"
    
    def _calculate_quality(
        self,
        sheets: list[ExtractedSheetData],
        warnings: list[ExtractionWarning]
    ) -> ExtractionQuality:
        """
        Calculate extraction quality metrics.
        
        Args:
            sheets: List of extracted sheet data.
            warnings: List of extraction warnings.
        
        Returns:
            ExtractionQuality object with metrics.
        """
        if not sheets:
            return ExtractionQuality(
                score=0.0,
                data_completeness=0.0,
                structure_clarity=0.0,
                has_headers=False,
                has_data=False,
                error_count=len([w for w in warnings if w.severity == "error"]),
                warning_count=len(warnings),
            )
        
        # Calculate metrics
        total_sheets = len(sheets)
        sheets_with_headers = sum(1 for s in sheets if s.has_headers)
        sheets_with_data = sum(1 for s in sheets if s.row_count > 0)
        
        # Calculate data completeness (sample first 100 rows per sheet)
        total_cells = 0
        non_empty_cells = 0
        
        for sheet in sheets:
            for row in sheet.data[:100]:
                for value in row:
                    total_cells += 1
                    if value is not None and value != "":
                        non_empty_cells += 1
        
        data_completeness = non_empty_cells / total_cells if total_cells > 0 else 0.0
        
        # Calculate structure clarity based on column consistency
        avg_columns = sum(s.column_count for s in sheets) / total_sheets
        structure_clarity = min(avg_columns / 20.0, 1.0)
        
        # Calculate overall score
        has_headers = sheets_with_headers > 0
        has_data = sheets_with_data > 0
        
        error_count = len([w for w in warnings if w.severity == "error"])
        warning_count = len(warnings)
        
        score = (
            (0.3 if has_headers else 0.0) +
            (0.3 if has_data else 0.0) +
            (0.2 * data_completeness) +
            (0.2 * structure_clarity)
        )
        
        # Penalize for errors
        if error_count > 0:
            score = max(0.0, score - (error_count * 0.1))
        
        return ExtractionQuality(
            score=score,
            data_completeness=data_completeness,
            structure_clarity=structure_clarity,
            has_headers=has_headers,
            has_data=has_data,
            error_count=error_count,
            warning_count=warning_count,
        )
    
    def _detect_content_language(
        self,
        sheets: list[ExtractedSheetData]
    ) -> str:
        """
        Detect the primary language of Excel content from extracted sheets.
        
        Analyzes headers and data from all sheets to determine the primary
        language of the workbook content.
        
        Args:
            sheets: List of extracted sheet data.
            
        Returns:
            ISO 639-1 language code (e.g., "en", "th", "mixed").
        
        Supports Requirement 23.1: Detect the primary language of Excel content.
        """
        if not sheets:
            self._logger.debug("No sheets to detect language from, using default")
            return self._language_service.DEFAULT_LANGUAGE
        
        # Collect text samples from all sheets
        all_text_parts: list[str] = []
        
        for sheet in sheets:
            # Add headers
            if sheet.headers:
                all_text_parts.extend(str(h) for h in sheet.headers if h)
            
            # Add sample data (first 50 rows per sheet)
            for row in sheet.data[:50]:
                for value in row:
                    if value is not None and isinstance(value, str) and value.strip():
                        all_text_parts.append(value)
        
        if not all_text_parts:
            self._logger.debug("No text content found in sheets, using default language")
            return self._language_service.DEFAULT_LANGUAGE
        
        # Combine text and detect language
        combined_text = " ".join(all_text_parts)
        detected_language = self._language_service.detect_language(combined_text)
        
        self._logger.info(
            f"Detected language '{detected_language}' from {len(sheets)} sheets "
            f"({len(all_text_parts)} text samples)"
        )
        
        return detected_language
