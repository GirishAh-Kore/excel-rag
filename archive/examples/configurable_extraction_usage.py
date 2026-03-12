"""
Example usage of configurable extraction with LLM summarization.

This script demonstrates:
1. Using different extraction strategies
2. LLM-based sheet summarization
3. Sheet ranking for disambiguation
"""

import asyncio
import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.extraction import (
    ConfigurableExtractor,
    ExtractionConfig,
    ExtractionStrategy,
)


def create_sample_excel_with_multiple_sheets() -> bytes:
    """Create a sample Excel file with multiple sheets."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: Sales Data
    ws1 = wb.create_sheet("Sales Data")
    ws1['A1'] = "Product"
    ws1['B1'] = "Region"
    ws1['C1'] = "Revenue"
    ws1['D1'] = "Date"
    
    ws1['A2'] = "Laptop"
    ws1['B2'] = "North"
    ws1['C2'] = 50000
    ws1['D2'] = datetime(2024, 1, 15)
    
    ws1['A3'] = "Mouse"
    ws1['B3'] = "South"
    ws1['C3'] = 1500
    ws1['D3'] = datetime(2024, 1, 16)
    
    # Sheet 2: Expense Report
    ws2 = wb.create_sheet("Expense Report")
    ws2['A1'] = "Category"
    ws2['B1'] = "Amount"
    ws2['C1'] = "Date"
    ws2['D1'] = "Approved"
    
    ws2['A2'] = "Travel"
    ws2['B2'] = 2500
    ws2['C2'] = datetime(2024, 1, 10)
    ws2['D2'] = "Yes"
    
    ws2['A3'] = "Marketing"
    ws2['B3'] = 5000
    ws2['C3'] = datetime(2024, 1, 12)
    ws2['D3'] = "Yes"
    
    # Sheet 3: Employee List
    ws3 = wb.create_sheet("Employee List")
    ws3['A1'] = "Name"
    ws3['B1'] = "Department"
    ws3['C1'] = "Salary"
    ws3['D1'] = "Start Date"
    
    ws3['A2'] = "John Doe"
    ws3['B2'] = "Engineering"
    ws3['C2'] = 80000
    ws3['D2'] = datetime(2023, 6, 1)
    
    ws3['A3'] = "Jane Smith"
    ws3['B3'] = "Marketing"
    ws3['C3'] = 75000
    ws3['D3'] = datetime(2023, 8, 15)
    
    # Save to bytes
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.read()


async def main():
    """Demonstrate configurable extraction."""
    print("=" * 70)
    print("Configurable Extraction with LLM Summarization Example")
    print("=" * 70)
    
    # Create sample Excel file
    print("\n1. Creating sample Excel file with multiple sheets...")
    file_content = create_sample_excel_with_multiple_sheets()
    print(f"   Created file with {len(file_content)} bytes")
    
    # Configure extraction
    print("\n2. Configuring extraction...")
    config = ExtractionConfig(
        default_strategy=ExtractionStrategy.OPENPYXL,
        max_rows_per_sheet=10000,
        max_file_size_mb=100,
        enable_llm_summarization=True,  # Enable LLM summarization
        summarization_provider="openai",
        summarization_model="gpt-4o-mini",  # Use cheaper model for summaries
        summarization_max_tokens=150,
        enable_gemini=False,  # Gemini not implemented yet
        gemini_api_key=None,
        gemini_model="gemini-1.5-flash",
        gemini_fallback_on_error=False,
        enable_llamaparse=False,  # LlamaParse not implemented yet
        llamaparse_api_key=None,
        use_auto_strategy=False,
        complexity_threshold=0.7
    )
    
    print(f"   Strategy: {config.default_strategy}")
    print(f"   LLM Summarization: {config.enable_llm_summarization}")
    print(f"   Summarization Provider: {config.summarization_provider}")
    
    # Initialize extractor
    print("\n3. Initializing ConfigurableExtractor...")
    extractor = ConfigurableExtractor(config)
    
    # Extract workbook
    print("\n4. Extracting workbook...")
    try:
        workbook_data = await extractor.extract_workbook(
            file_content=file_content,
            file_id="sample_multi_123",
            file_name="company_data.xlsx",
            file_path="/examples/company_data.xlsx",
            modified_time=datetime.now(),
            strategy=ExtractionStrategy.OPENPYXL
        )
        
        print(f"   ✓ Successfully extracted workbook")
        print(f"   - File: {workbook_data.file_name}")
        print(f"   - Sheets: {len(workbook_data.sheets)}")
        
    except Exception as e:
        print(f"   ✗ Extraction failed: {e}")
        return
    
    # Display sheet information with LLM summaries
    print("\n5. Sheet Details with LLM Summaries:")
    for idx, sheet in enumerate(workbook_data.sheets, start=1):
        print(f"\n   Sheet {idx}: {sheet.sheet_name}")
        print(f"   - Headers: {', '.join(sheet.headers)}")
        print(f"   - Rows: {sheet.row_count}")
        print(f"   - Columns: {sheet.column_count}")
        
        # Show basic summary
        print(f"\n   Basic Summary:")
        print(f"   {sheet.summary}")
        
        # Show LLM summary if available
        if sheet.llm_summary:
            print(f"\n   LLM Summary:")
            print(f"   {sheet.llm_summary}")
            print(f"   (Generated at: {sheet.summary_generated_at})")
        else:
            print(f"\n   LLM Summary: Not generated (check API key configuration)")
    
    # Demonstrate sheet ranking for disambiguation
    print("\n6. Demonstrating Sheet Ranking for Query:")
    query = "What were the expenses last month?"
    print(f"   Query: \"{query}\"")
    
    if extractor.summarizer:
        print("\n   Ranking sheets by relevance...")
        candidate_sheets = [(sheet, workbook_data.file_name) for sheet in workbook_data.sheets]
        
        try:
            ranked_sheets = await extractor.summarizer.rank_sheets_for_query(
                query=query,
                candidate_sheets=candidate_sheets,
                use_llm_ranking=True
            )
            
            print("\n   Ranked Results:")
            for idx, (sheet, score, summary) in enumerate(ranked_sheets, start=1):
                print(f"\n   {idx}. {sheet.sheet_name} (Score: {score:.2f})")
                print(f"      {summary[:150]}...")
        
        except Exception as e:
            print(f"   ✗ Ranking failed: {e}")
            print(f"      (This requires a valid API key for the LLM provider)")
    else:
        print("   Summarizer not available (LLM summarization disabled)")
    
    # Show extraction statistics
    print("\n7. Extraction Statistics:")
    failed_files = extractor.get_failed_files()
    if failed_files:
        print(f"   Failed files: {len(failed_files)}")
        for failed in failed_files:
            print(f"   - {failed['file_name']}: {failed['error_type']}")
    else:
        print("   No failed files")
    
    print("\n" + "=" * 70)
    print("Extraction Complete!")
    print("=" * 70)
    
    print("\nNote: To enable LLM summarization, set these environment variables:")
    print("  - ENABLE_LLM_SUMMARIZATION=true")
    print("  - SUMMARIZATION_PROVIDER=openai (or anthropic, gemini)")
    print("  - LLM_API_KEY=your_api_key")
    print("\nFor Gemini extraction (optional):")
    print("  - ENABLE_GEMINI_EXTRACTION=true")
    print("  - GEMINI_API_KEY=your_gemini_api_key")


if __name__ == "__main__":
    asyncio.run(main())
