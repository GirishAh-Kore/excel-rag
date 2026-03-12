"""
Export Service Module.

This module implements export capabilities for query results and chunk
data to various formats including CSV, Excel (.xlsx), and JSON.

Key Features:
- Support export to CSV, Excel (.xlsx), JSON formats
- Preserve data types and formatting for Excel export
- Support scheduled exports for recurring reports

Supports Requirements 26.1, 26.2, 26.3, 26.4, 26.5.
"""

import csv
import io
import json
import logging
import uuid
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import Any, Optional, Protocol, Union, runtime_checkable

from src.exceptions import ExportError

logger = logging.getLogger(__name__)


# =============================================================================
# Constants
# =============================================================================

# Maximum rows for export without streaming
MAX_EXPORT_ROWS = 100000

# Default export filename prefix
DEFAULT_FILENAME_PREFIX = "export"

# Supported date formats for Excel
EXCEL_DATE_FORMAT = "YYYY-MM-DD HH:MM:SS"


# =============================================================================
# Enums
# =============================================================================


class ExportFormat(str, Enum):
    """
    Supported export formats.
    
    Supports Requirement 26.1: Support export to CSV, Excel (.xlsx), JSON formats.
    """
    CSV = "csv"
    EXCEL = "xlsx"
    JSON = "json"


class ScheduleFrequency(str, Enum):
    """Frequency options for scheduled exports."""
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"


# =============================================================================
# Configuration
# =============================================================================


@dataclass
class ExportServiceConfig:
    """
    Configuration for ExportService.
    
    Attributes:
        max_rows: Maximum rows allowed in a single export.
        default_format: Default export format if not specified.
        include_metadata: Whether to include metadata in exports.
        excel_date_format: Date format for Excel exports.
    """
    max_rows: int = MAX_EXPORT_ROWS
    default_format: ExportFormat = ExportFormat.CSV
    include_metadata: bool = True
    excel_date_format: str = EXCEL_DATE_FORMAT
    
    def __post_init__(self) -> None:
        """Validate configuration values."""
        if self.max_rows <= 0:
            raise ValueError(f"max_rows must be positive, got {self.max_rows}")


# =============================================================================
# Data Classes
# =============================================================================


@dataclass
class ScheduledExport:
    """
    Scheduled export configuration.
    
    Attributes:
        schedule_id: Unique identifier for the schedule.
        name: Human-readable name for the schedule.
        query_or_result_id: Query text or result ID to export.
        export_format: Format for the export.
        frequency: How often to run the export.
        created_by: User who created the schedule.
        created_at: When the schedule was created.
        last_run_at: When the export last ran.
        next_run_at: When the export will next run.
        is_active: Whether the schedule is active.
        destination: Optional destination path or URL.
    """
    schedule_id: str
    name: str
    query_or_result_id: str
    export_format: ExportFormat
    frequency: ScheduleFrequency
    created_by: str
    created_at: datetime
    last_run_at: Optional[datetime] = None
    next_run_at: Optional[datetime] = None
    is_active: bool = True
    destination: Optional[str] = None
    
    def to_dict(self) -> dict[str, Any]:
        """Convert to dictionary for serialization."""
        return {
            "schedule_id": self.schedule_id,
            "name": self.name,
            "query_or_result_id": self.query_or_result_id,
            "export_format": self.export_format.value,
            "frequency": self.frequency.value,
            "created_by": self.created_by,
            "created_at": self.created_at.isoformat(),
            "last_run_at": self.last_run_at.isoformat() if self.last_run_at else None,
            "next_run_at": self.next_run_at.isoformat() if self.next_run_at else None,
            "is_active": self.is_active,
            "destination": self.destination,
        }


@dataclass
class ExportResult:
    """
    Result of an export operation.
    
    Attributes:
        export_id: Unique identifier for the export.
        format: Format of the exported data.
        data: Exported data as bytes.
        filename: Suggested filename for the export.
        row_count: Number of rows exported.
        created_at: When the export was created.
        metadata: Optional metadata about the export.
    """
    export_id: str
    format: ExportFormat
    data: bytes
    filename: str
    row_count: int
    created_at: datetime
    metadata: Optional[dict[str, Any]] = None


# =============================================================================
# Protocols
# =============================================================================


@runtime_checkable
class ExportStoreProtocol(Protocol):
    """
    Protocol for scheduled export storage.
    
    Implementations must provide methods for CRUD operations on
    scheduled exports.
    """
    
    def create_schedule(self, schedule: ScheduledExport) -> bool:
        """Create a new scheduled export."""
        ...
    
    def get_schedule(self, schedule_id: str) -> Optional[ScheduledExport]:
        """Get a scheduled export by ID."""
        ...
    
    def get_schedules_for_user(self, user_id: str) -> list[ScheduledExport]:
        """Get all scheduled exports for a user."""
        ...
    
    def get_due_schedules(self, as_of: datetime) -> list[ScheduledExport]:
        """Get all schedules due for execution."""
        ...
    
    def update_schedule(self, schedule: ScheduledExport) -> bool:
        """Update a scheduled export."""
        ...
    
    def delete_schedule(self, schedule_id: str) -> bool:
        """Delete a scheduled export."""
        ...


# =============================================================================
# Export Service
# =============================================================================


class ExportService:
    """
    Service for exporting query results and chunk data.
    
    Provides functionality for exporting data to CSV, Excel (.xlsx),
    and JSON formats with support for scheduled recurring exports.
    
    All dependencies are injected via constructor following DIP.
    
    Implements Requirements:
    - 26.1: Support export to CSV, Excel (.xlsx), JSON formats
    - 26.2: Support exporting chunk listings and metadata
    - 26.3: Expose POST endpoint for exports (via API layer)
    - 26.4: Preserve data types and formatting for Excel export
    - 26.5: Support scheduled exports for recurring reports
    
    Example:
        >>> service = ExportService(
        ...     export_store=store,
        ...     config=ExportServiceConfig()
        ... )
        >>> result = service.export_data(
        ...     data=query_results,
        ...     format=ExportFormat.EXCEL,
        ...     filename_prefix="sales_report"
        ... )
    """
    
    def __init__(
        self,
        export_store: ExportStoreProtocol,
        config: Optional[ExportServiceConfig] = None
    ) -> None:
        """
        Initialize ExportService with injected dependencies.
        
        Args:
            export_store: Service for storing scheduled exports.
            config: Optional configuration (uses defaults if not provided).
            
        Raises:
            ValueError: If export_store is None.
        """
        if export_store is None:
            raise ValueError("export_store is required")
        
        self._export_store = export_store
        self._config = config or ExportServiceConfig()
        
        logger.info(
            f"ExportService initialized with max_rows={self._config.max_rows}"
        )

    def export_data(
        self,
        data: list[dict[str, Any]],
        format: Optional[ExportFormat] = None,
        filename_prefix: Optional[str] = None,
        include_metadata: Optional[bool] = None
    ) -> ExportResult:
        """
        Export data to the specified format.
        
        Converts the provided data to the requested format and returns
        the result as bytes with suggested filename.
        
        Args:
            data: List of dictionaries to export.
            format: Export format (defaults to config default).
            filename_prefix: Prefix for the filename.
            include_metadata: Whether to include metadata.
            
        Returns:
            ExportResult with exported data and metadata.
            
        Raises:
            ExportError: If export fails or data is invalid.
        """
        # Use defaults from config
        export_format = format or self._config.default_format
        prefix = filename_prefix or DEFAULT_FILENAME_PREFIX
        include_meta = include_metadata if include_metadata is not None else self._config.include_metadata
        
        # Validate data
        if not data:
            raise ExportError(
                "Cannot export empty data",
                details={"format": export_format.value}
            )
        
        if len(data) > self._config.max_rows:
            raise ExportError(
                f"Data exceeds maximum export size of {self._config.max_rows} rows",
                details={
                    "row_count": len(data),
                    "max_rows": self._config.max_rows
                }
            )
        
        # Generate export ID and timestamp
        export_id = f"exp_{uuid.uuid4().hex[:16]}"
        timestamp = datetime.utcnow()
        timestamp_str = timestamp.strftime("%Y%m%d_%H%M%S")
        
        # Export based on format
        try:
            if export_format == ExportFormat.CSV:
                exported_data, filename = self._export_to_csv(
                    data=data,
                    prefix=prefix,
                    timestamp_str=timestamp_str
                )
            elif export_format == ExportFormat.EXCEL:
                exported_data, filename = self._export_to_excel(
                    data=data,
                    prefix=prefix,
                    timestamp_str=timestamp_str
                )
            elif export_format == ExportFormat.JSON:
                exported_data, filename = self._export_to_json(
                    data=data,
                    prefix=prefix,
                    timestamp_str=timestamp_str,
                    include_metadata=include_meta
                )
            else:
                raise ExportError(
                    f"Unsupported export format: {export_format}",
                    details={"format": export_format.value}
                )
            
            logger.info(
                f"Exported {len(data)} rows to {export_format.value} "
                f"(export_id={export_id})"
            )
            
            # Build metadata
            metadata = None
            if include_meta:
                metadata = {
                    "export_id": export_id,
                    "format": export_format.value,
                    "row_count": len(data),
                    "column_count": len(data[0]) if data else 0,
                    "columns": list(data[0].keys()) if data else [],
                    "exported_at": timestamp.isoformat(),
                }
            
            return ExportResult(
                export_id=export_id,
                format=export_format,
                data=exported_data,
                filename=filename,
                row_count=len(data),
                created_at=timestamp,
                metadata=metadata
            )
            
        except ExportError:
            raise
        except Exception as e:
            logger.error(f"Export failed: {e}", exc_info=True)
            raise ExportError(
                f"Export failed: {str(e)}",
                details={
                    "format": export_format.value,
                    "row_count": len(data)
                }
            )
    
    def _export_to_csv(
        self,
        data: list[dict[str, Any]],
        prefix: str,
        timestamp_str: str
    ) -> tuple[bytes, str]:
        """
        Export data to CSV format.
        
        Args:
            data: Data to export.
            prefix: Filename prefix.
            timestamp_str: Timestamp string for filename.
            
        Returns:
            Tuple of (csv_bytes, filename).
        """
        output = io.StringIO()
        
        # Get all unique keys from data
        fieldnames = self._get_all_fieldnames(data)
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for row in data:
            # Convert complex types to strings
            processed_row = self._process_row_for_csv(row)
            writer.writerow(processed_row)
        
        csv_content = output.getvalue()
        filename = f"{prefix}_{timestamp_str}.csv"
        
        return csv_content.encode('utf-8'), filename
    
    def _export_to_excel(
        self,
        data: list[dict[str, Any]],
        prefix: str,
        timestamp_str: str
    ) -> tuple[bytes, str]:
        """
        Export data to Excel (.xlsx) format with preserved data types.
        
        Supports Requirement 26.4: Preserve data types and formatting for Excel export.
        
        Args:
            data: Data to export.
            prefix: Filename prefix.
            timestamp_str: Timestamp string for filename.
            
        Returns:
            Tuple of (excel_bytes, filename).
            
        Raises:
            ExportError: If openpyxl is not available.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ExportError(
                "openpyxl is required for Excel export",
                details={"format": "xlsx"}
            )
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # Get all unique keys from data
        fieldnames = self._get_all_fieldnames(data)
        
        # Write header row with styling
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCCCCC",
            end_color="CCCCCC",
            fill_type="solid"
        )
        
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            cell = ws.cell(row=1, column=col_idx, value=fieldname)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data rows with type preservation
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, fieldname in enumerate(fieldnames, start=1):
                value = row_data.get(fieldname)
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Preserve data types
                if value is None:
                    cell.value = None
                elif isinstance(value, datetime):
                    cell.value = value
                    cell.number_format = self._config.excel_date_format
                elif isinstance(value, bool):
                    cell.value = value
                elif isinstance(value, (int, float)):
                    cell.value = value
                    # Format currency if detected
                    if self._is_currency_value(fieldname, value):
                        cell.number_format = '"$"#,##0.00'
                    elif self._is_percentage_value(fieldname, value):
                        cell.number_format = '0.00%'
                elif isinstance(value, (list, dict)):
                    cell.value = json.dumps(value)
                else:
                    cell.value = str(value)
        
        # Auto-adjust column widths
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            column_letter = get_column_letter(col_idx)
            max_length = len(str(fieldname))
            
            for row_idx in range(2, min(len(data) + 2, 100)):  # Sample first 100 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{prefix}_{timestamp_str}.xlsx"
        
        return output.getvalue(), filename
    
    def _export_to_json(
        self,
        data: list[dict[str, Any]],
        prefix: str,
        timestamp_str: str,
        include_metadata: bool
    ) -> tuple[bytes, str]:
        """
        Export data to JSON format.
        
        Args:
            data: Data to export.
            prefix: Filename prefix.
            timestamp_str: Timestamp string for filename.
            include_metadata: Whether to include metadata wrapper.
            
        Returns:
            Tuple of (json_bytes, filename).
        """
        if include_metadata:
            export_obj = {
                "metadata": {
                    "exported_at": datetime.utcnow().isoformat(),
                    "row_count": len(data),
                    "column_count": len(data[0]) if data else 0,
                },
                "data": data
            }
        else:
            export_obj = data
        
        json_content = json.dumps(
            export_obj,
            indent=2,
            default=self._json_serializer
        )
        
        filename = f"{prefix}_{timestamp_str}.json"
        
        return json_content.encode('utf-8'), filename
    
    def _get_all_fieldnames(self, data: list[dict[str, Any]]) -> list[str]:
        """
        Get all unique field names from data, preserving order.
        
        Args:
            data: List of dictionaries.
            
        Returns:
            List of unique field names.
        """
        seen = set()
        fieldnames = []
        
        for row in data:
            for key in row.keys():
                if key not in seen:
                    seen.add(key)
                    fieldnames.append(key)
        
        return fieldnames
    
    def _process_row_for_csv(self, row: dict[str, Any]) -> dict[str, str]:
        """
        Process a row for CSV export, converting complex types.
        
        Args:
            row: Row data dictionary.
            
        Returns:
            Processed row with string values.
        """
        processed = {}
        
        for key, value in row.items():
            if value is None:
                processed[key] = ""
            elif isinstance(value, datetime):
                processed[key] = value.isoformat()
            elif isinstance(value, (list, dict)):
                processed[key] = json.dumps(value)
            elif isinstance(value, bool):
                processed[key] = str(value).lower()
            else:
                processed[key] = str(value)
        
        return processed
    
    def _is_currency_value(self, fieldname: str, value: Any) -> bool:
        """Check if a value appears to be currency based on field name."""
        currency_indicators = [
            "price", "cost", "amount", "revenue", "sales",
            "total", "fee", "payment", "balance", "budget"
        ]
        fieldname_lower = fieldname.lower()
        return any(ind in fieldname_lower for ind in currency_indicators)
    
    def _is_percentage_value(self, fieldname: str, value: Any) -> bool:
        """Check if a value appears to be a percentage based on field name."""
        percentage_indicators = [
            "percent", "pct", "rate", "ratio", "growth",
            "margin", "yield", "return"
        ]
        fieldname_lower = fieldname.lower()
        return any(ind in fieldname_lower for ind in percentage_indicators)
    
    @staticmethod
    def _json_serializer(obj: Any) -> Any:
        """Custom JSON serializer for non-standard types."""
        if isinstance(obj, datetime):
            return obj.isoformat()
        if hasattr(obj, 'to_dict'):
            return obj.to_dict()
        if hasattr(obj, '__dict__'):
            return obj.__dict__
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

    # =========================================================================
    # Scheduled Export Methods
    # =========================================================================
    
    def create_scheduled_export(
        self,
        name: str,
        query_or_result_id: str,
        export_format: ExportFormat,
        frequency: ScheduleFrequency,
        created_by: str,
        destination: Optional[str] = None
    ) -> ScheduledExport:
        """
        Create a scheduled export for recurring reports.
        
        Supports Requirement 26.5: Support scheduled exports for recurring reports.
        
        Args:
            name: Human-readable name for the schedule.
            query_or_result_id: Query text or result ID to export.
            export_format: Format for the export.
            frequency: How often to run the export.
            created_by: User creating the schedule.
            destination: Optional destination path or URL.
            
        Returns:
            Created ScheduledExport.
            
        Raises:
            ExportError: If schedule creation fails.
        """
        # Validate inputs
        if not name or not name.strip():
            raise ExportError(
                "Schedule name cannot be empty",
                details={"name": name}
            )
        
        if not query_or_result_id or not query_or_result_id.strip():
            raise ExportError(
                "Query or result ID cannot be empty",
                details={"query_or_result_id": query_or_result_id}
            )
        
        if not created_by or not created_by.strip():
            raise ExportError(
                "created_by cannot be empty",
                details={"created_by": created_by}
            )
        
        # Generate schedule ID
        schedule_id = f"sched_{uuid.uuid4().hex[:16]}"
        now = datetime.utcnow()
        
        # Calculate next run time
        next_run = self._calculate_next_run(now, frequency)
        
        schedule = ScheduledExport(
            schedule_id=schedule_id,
            name=name.strip(),
            query_or_result_id=query_or_result_id,
            export_format=export_format,
            frequency=frequency,
            created_by=created_by,
            created_at=now,
            next_run_at=next_run,
            is_active=True,
            destination=destination
        )
        
        # Store schedule
        success = self._export_store.create_schedule(schedule)
        if not success:
            raise ExportError(
                "Failed to create scheduled export",
                details={"schedule_id": schedule_id, "name": name}
            )
        
        logger.info(
            f"Created scheduled export {schedule_id}: '{name}' "
            f"({frequency.value}, {export_format.value})"
        )
        
        return schedule
    
    def get_scheduled_export(self, schedule_id: str) -> Optional[ScheduledExport]:
        """
        Get a scheduled export by ID.
        
        Args:
            schedule_id: Unique schedule identifier.
            
        Returns:
            ScheduledExport if found, None otherwise.
        """
        return self._export_store.get_schedule(schedule_id)
    
    def get_scheduled_exports_for_user(self, user_id: str) -> list[ScheduledExport]:
        """
        Get all scheduled exports for a user.
        
        Args:
            user_id: User ID to get schedules for.
            
        Returns:
            List of ScheduledExport objects.
        """
        return self._export_store.get_schedules_for_user(user_id)
    
    def update_scheduled_export(
        self,
        schedule_id: str,
        name: Optional[str] = None,
        export_format: Optional[ExportFormat] = None,
        frequency: Optional[ScheduleFrequency] = None,
        is_active: Optional[bool] = None,
        destination: Optional[str] = None,
        user_id: Optional[str] = None
    ) -> ScheduledExport:
        """
        Update a scheduled export.
        
        Args:
            schedule_id: ID of the schedule to update.
            name: Optional new name.
            export_format: Optional new format.
            frequency: Optional new frequency.
            is_active: Optional new active status.
            destination: Optional new destination.
            user_id: User ID for authorization check.
            
        Returns:
            Updated ScheduledExport.
            
        Raises:
            ExportError: If schedule not found or update fails.
        """
        # Get existing schedule
        schedule = self._export_store.get_schedule(schedule_id)
        if schedule is None:
            raise ExportError(
                f"Scheduled export not found: {schedule_id}",
                details={"schedule_id": schedule_id}
            )
        
        # Check authorization (only creator can update)
        if user_id and schedule.created_by != user_id:
            raise ExportError(
                "Not authorized to update this scheduled export",
                details={
                    "schedule_id": schedule_id,
                    "user_id": user_id,
                    "created_by": schedule.created_by
                }
            )
        
        # Update fields
        updated_name = name.strip() if name else schedule.name
        updated_format = export_format if export_format else schedule.export_format
        updated_frequency = frequency if frequency else schedule.frequency
        updated_active = is_active if is_active is not None else schedule.is_active
        updated_destination = destination if destination is not None else schedule.destination
        
        # Recalculate next run if frequency changed
        next_run = schedule.next_run_at
        if frequency and frequency != schedule.frequency:
            next_run = self._calculate_next_run(datetime.utcnow(), frequency)
        
        updated_schedule = ScheduledExport(
            schedule_id=schedule.schedule_id,
            name=updated_name,
            query_or_result_id=schedule.query_or_result_id,
            export_format=updated_format,
            frequency=updated_frequency,
            created_by=schedule.created_by,
            created_at=schedule.created_at,
            last_run_at=schedule.last_run_at,
            next_run_at=next_run,
            is_active=updated_active,
            destination=updated_destination
        )
        
        # Store updated schedule
        success = self._export_store.update_schedule(updated_schedule)
        if not success:
            raise ExportError(
                "Failed to update scheduled export",
                details={"schedule_id": schedule_id}
            )
        
        logger.info(f"Updated scheduled export {schedule_id}")
        return updated_schedule
    
    def delete_scheduled_export(
        self,
        schedule_id: str,
        user_id: Optional[str] = None
    ) -> bool:
        """
        Delete a scheduled export.
        
        Args:
            schedule_id: ID of the schedule to delete.
            user_id: User ID for authorization check.
            
        Returns:
            True if deleted successfully.
            
        Raises:
            ExportError: If schedule not found or not authorized.
        """
        # Get existing schedule for authorization check
        schedule = self._export_store.get_schedule(schedule_id)
        if schedule is None:
            raise ExportError(
                f"Scheduled export not found: {schedule_id}",
                details={"schedule_id": schedule_id}
            )
        
        # Check authorization (only creator can delete)
        if user_id and schedule.created_by != user_id:
            raise ExportError(
                "Not authorized to delete this scheduled export",
                details={
                    "schedule_id": schedule_id,
                    "user_id": user_id,
                    "created_by": schedule.created_by
                }
            )
        
        success = self._export_store.delete_schedule(schedule_id)
        if success:
            logger.info(f"Deleted scheduled export {schedule_id}")
        
        return success
    
    def get_due_schedules(self) -> list[ScheduledExport]:
        """
        Get all scheduled exports that are due for execution.
        
        Returns:
            List of ScheduledExport objects due for execution.
        """
        return self._export_store.get_due_schedules(datetime.utcnow())
    
    def mark_schedule_executed(
        self,
        schedule_id: str
    ) -> ScheduledExport:
        """
        Mark a scheduled export as executed and update next run time.
        
        Args:
            schedule_id: ID of the schedule that was executed.
            
        Returns:
            Updated ScheduledExport with new next_run_at.
            
        Raises:
            ExportError: If schedule not found or update fails.
        """
        schedule = self._export_store.get_schedule(schedule_id)
        if schedule is None:
            raise ExportError(
                f"Scheduled export not found: {schedule_id}",
                details={"schedule_id": schedule_id}
            )
        
        now = datetime.utcnow()
        next_run = self._calculate_next_run(now, schedule.frequency)
        
        updated_schedule = ScheduledExport(
            schedule_id=schedule.schedule_id,
            name=schedule.name,
            query_or_result_id=schedule.query_or_result_id,
            export_format=schedule.export_format,
            frequency=schedule.frequency,
            created_by=schedule.created_by,
            created_at=schedule.created_at,
            last_run_at=now,
            next_run_at=next_run,
            is_active=schedule.is_active,
            destination=schedule.destination
        )
        
        success = self._export_store.update_schedule(updated_schedule)
        if not success:
            raise ExportError(
                "Failed to update schedule after execution",
                details={"schedule_id": schedule_id}
            )
        
        logger.info(
            f"Marked scheduled export {schedule_id} as executed, "
            f"next run at {next_run.isoformat()}"
        )
        
        return updated_schedule
    
    def _calculate_next_run(
        self,
        from_time: datetime,
        frequency: ScheduleFrequency
    ) -> datetime:
        """
        Calculate the next run time based on frequency.
        
        Args:
            from_time: Starting time for calculation.
            frequency: Schedule frequency.
            
        Returns:
            Next run datetime.
        """
        from datetime import timedelta
        
        if frequency == ScheduleFrequency.DAILY:
            return from_time + timedelta(days=1)
        elif frequency == ScheduleFrequency.WEEKLY:
            return from_time + timedelta(weeks=1)
        elif frequency == ScheduleFrequency.MONTHLY:
            # Add approximately one month (30 days)
            return from_time + timedelta(days=30)
        else:
            # Default to daily
            return from_time + timedelta(days=1)
