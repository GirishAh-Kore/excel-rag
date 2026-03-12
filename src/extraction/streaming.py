"""
Streaming extraction for large Excel files.

This module provides streaming extraction capabilities for handling large
Excel files (>100MB) and sheets with >1M rows. It implements chunked
processing with configurable memory limits and graceful degradation.

Supports Requirements:
- 40.1: Support files larger than 100MB
- 40.2: Implement chunked processing for sheets >1M rows
- 40.5: Configure memory limits with graceful degradation
"""

import logging
import os
from abc import ABC, abstractmethod
from dataclasses import dataclass, field
from typing import Any, Generator, Iterator, Optional, Protocol

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.exceptions import ExtractionError
from src.models.excel_features import ExtractedSheetData, ExtractionQuality


logger = logging.getLogger(__name__)


# =============================================================================
# Configuration
# =============================================================================

# Default configuration constants
DEFAULT_CHUNK_SIZE_ROWS = 10000
DEFAULT_MAX_MEMORY_MB = 512
DEFAULT_LARGE_FILE_THRESHOLD_MB = 100
DEFAULT_LARGE_SHEET_THRESHOLD_ROWS = 1_000_000


@dataclass
class StreamingExtractionConfig:
    """
    Configuration for streaming extraction.
    
    Attributes:
        chunk_size_rows: Number of rows to process per chunk.
        max_memory_mb: Maximum memory usage in MB before degradation.
        large_file_threshold_mb: File size threshold for streaming mode.
        large_sheet_threshold_rows: Row count threshold for chunked processing.
        enable_memory_monitoring: Whether to monitor memory usage.
        graceful_degradation: Whether to reduce quality on memory pressure.
    """
    chunk_size_rows: int = DEFAULT_CHUNK_SIZE_ROWS
    max_memory_mb: int = DEFAULT_MAX_MEMORY_MB
    large_file_threshold_mb: int = DEFAULT_LARGE_FILE_THRESHOLD_MB
    large_sheet_threshold_rows: int = DEFAULT_LARGE_SHEET_THRESHOLD_ROWS
    enable_memory_monitoring: bool = True
    graceful_degradation: bool = True
    
    def __post_init__(self) -> None:
        """Validate configuration values."""
        if self.chunk_size_rows < 100:
            raise ValueError(
                f"chunk_size_rows must be at least 100, got {self.chunk_size_rows}"
            )
        if self.max_memory_mb < 64:
            raise ValueError(
                f"max_memory_mb must be at least 64, got {self.max_memory_mb}"
            )
        if self.large_file_threshold_mb < 1:
            raise ValueError(
                f"large_file_threshold_mb must be at least 1, got {self.large_file_threshold_mb}"
            )


# =============================================================================
# Data Models
# =============================================================================

@dataclass
class StreamingChunk:
    """
    A chunk of data from streaming extraction.
    
    Attributes:
        sheet_name: Name of the source sheet.
        start_row: Starting row index (1-based).
        end_row: Ending row index (1-based, inclusive).
        headers: Column headers (from first chunk or detected).
        data: List of rows, where each row is a list of cell values.
        is_first_chunk: Whether this is the first chunk of the sheet.
        is_last_chunk: Whether this is the last chunk of the sheet.
        chunk_index: Index of this chunk (0-based).
        total_chunks: Total number of chunks (if known).
    """
    sheet_name: str
    start_row: int
    end_row: int
    headers: list[str]
    data: list[list[Any]]
    is_first_chunk: bool
    is_last_chunk: bool
    chunk_index: int
    total_chunks: Optional[int] = None
    
    @property
    def row_count(self) -> int:
        """Get the number of data rows in this chunk."""
        return len(self.data)


@dataclass
class StreamingExtractionResult:
    """
    Result of streaming extraction for a single file.
    
    Attributes:
        file_path: Path to the extracted file.
        file_size_mb: File size in megabytes.
        sheet_names: List of sheet names in the file.
        total_rows: Total rows across all sheets.
        chunks_processed: Number of chunks processed.
        used_streaming: Whether streaming mode was used.
        memory_warnings: List of memory-related warnings.
        quality: Extraction quality metrics.
    """
    file_path: str
    file_size_mb: float
    sheet_names: list[str]
    total_rows: int
    chunks_processed: int
    used_streaming: bool
    memory_warnings: list[str] = field(default_factory=list)
    quality: Optional[ExtractionQuality] = None


@dataclass
class MemoryStatus:
    """
    Current memory usage status.
    
    Attributes:
        current_mb: Current memory usage in MB.
        max_mb: Maximum allowed memory in MB.
        usage_percent: Memory usage as percentage of max.
        is_under_pressure: Whether memory is under pressure.
    """
    current_mb: float
    max_mb: float
    usage_percent: float
    is_under_pressure: bool


# =============================================================================
# Protocols
# =============================================================================

class MemoryMonitorProtocol(Protocol):
    """Protocol for memory monitoring services."""
    
    def get_current_usage_mb(self) -> float:
        """Get current memory usage in megabytes."""
        ...
    
    def is_under_pressure(self, threshold_mb: float) -> bool:
        """Check if memory usage exceeds threshold."""
        ...


class ChunkProcessorProtocol(Protocol):
    """Protocol for processing extracted chunks."""
    
    def process_chunk(self, chunk: StreamingChunk) -> None:
        """Process a single chunk of extracted data."""
        ...


# =============================================================================
# Memory Monitor Implementation
# =============================================================================

class ProcessMemoryMonitor:
    """
    Monitors process memory usage using psutil.
    
    Provides memory usage information for the current process to enable
    graceful degradation when memory limits are approached.
    """
    
    def __init__(self) -> None:
        """Initialize the memory monitor."""
        self._logger = logging.getLogger(__name__)
        self._psutil_available = self._check_psutil()
    
    def _check_psutil(self) -> bool:
        """Check if psutil is available."""
        try:
            import psutil
            return True
        except ImportError:
            self._logger.warning(
                "psutil not available, memory monitoring disabled"
            )
            return False
    
    def get_current_usage_mb(self) -> float:
        """
        Get current process memory usage in megabytes.
        
        Returns:
            Memory usage in MB, or 0.0 if psutil unavailable.
        """
        if not self._psutil_available:
            return 0.0
        
        try:
            import psutil
            process = psutil.Process()
            memory_info = process.memory_info()
            return memory_info.rss / (1024 * 1024)
        except Exception as e:
            self._logger.warning(f"Failed to get memory usage: {e}")
            return 0.0
    
    def is_under_pressure(self, threshold_mb: float) -> bool:
        """
        Check if memory usage exceeds the threshold.
        
        Args:
            threshold_mb: Memory threshold in megabytes.
            
        Returns:
            True if current usage exceeds threshold.
        """
        current = self.get_current_usage_mb()
        return current > threshold_mb
    
    def get_status(self, max_mb: float) -> MemoryStatus:
        """
        Get detailed memory status.
        
        Args:
            max_mb: Maximum allowed memory in MB.
            
        Returns:
            MemoryStatus with current usage details.
        """
        current = self.get_current_usage_mb()
        usage_percent = (current / max_mb * 100) if max_mb > 0 else 0.0
        
        return MemoryStatus(
            current_mb=current,
            max_mb=max_mb,
            usage_percent=usage_percent,
            is_under_pressure=current > max_mb * 0.8  # 80% threshold
        )


# =============================================================================
# Streaming Extractor
# =============================================================================

class StreamingExcelExtractor:
    """
    Streaming extractor for large Excel files.
    
    Processes Excel files in chunks to handle files larger than 100MB
    and sheets with more than 1M rows. Implements memory monitoring
    and graceful degradation when memory limits are approached.
    
    All dependencies are injected via constructor following DIP.
    
    Attributes:
        _config: Streaming extraction configuration.
        _memory_monitor: Memory monitoring service.
        _logger: Logger instance.
    
    Example:
        >>> config = StreamingExtractionConfig(chunk_size_rows=5000)
        >>> monitor = ProcessMemoryMonitor()
        >>> extractor = StreamingExcelExtractor(config, monitor)
        >>> for chunk in extractor.extract_streaming("large_file.xlsx"):
        ...     process_chunk(chunk)
    
    Supports Requirements 40.1, 40.2, 40.5.
    """
    
    def __init__(
        self,
        config: Optional[StreamingExtractionConfig] = None,
        memory_monitor: Optional[MemoryMonitorProtocol] = None
    ) -> None:
        """
        Initialize the streaming extractor.
        
        Args:
            config: Streaming extraction configuration.
            memory_monitor: Memory monitoring service.
        """
        self._config = config or StreamingExtractionConfig()
        self._memory_monitor = memory_monitor or ProcessMemoryMonitor()
        self._logger = logging.getLogger(__name__)
    
    def should_use_streaming(self, file_path: str) -> bool:
        """
        Determine if streaming extraction should be used for a file.
        
        Args:
            file_path: Path to the Excel file.
            
        Returns:
            True if file size exceeds the streaming threshold.
        
        Supports Requirement 40.1: Support files larger than 100MB.
        """
        try:
            file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
            return file_size_mb >= self._config.large_file_threshold_mb
        except OSError as e:
            self._logger.warning(f"Failed to get file size for {file_path}: {e}")
            return False
    
    def get_file_size_mb(self, file_path: str) -> float:
        """
        Get file size in megabytes.
        
        Args:
            file_path: Path to the file.
            
        Returns:
            File size in MB.
        """
        try:
            return os.path.getsize(file_path) / (1024 * 1024)
        except OSError:
            return 0.0
    
    def extract_streaming(
        self,
        file_path: str,
        sheet_names: Optional[list[str]] = None
    ) -> Generator[StreamingChunk, None, StreamingExtractionResult]:
        """
        Extract data from an Excel file using streaming.
        
        Yields chunks of data as they are extracted, allowing processing
        of large files without loading everything into memory.
        
        Args:
            file_path: Path to the Excel file.
            sheet_names: Optional list of specific sheets to extract.
                If None, extracts all sheets.
        
        Yields:
            StreamingChunk objects containing extracted data.
            
        Returns:
            StreamingExtractionResult with extraction summary.
        
        Raises:
            ExtractionError: If file cannot be read or is corrupted.
        
        Supports Requirements 40.1, 40.2, 40.5.
        """
        file_size_mb = self.get_file_size_mb(file_path)
        memory_warnings: list[str] = []
        total_rows = 0
        chunks_processed = 0
        all_sheet_names: list[str] = []
        
        self._logger.info(
            f"Starting streaming extraction for {file_path} "
            f"({file_size_mb:.1f} MB)"
        )
        
        try:
            # Load workbook in read-only mode for memory efficiency
            workbook = openpyxl.load_workbook(
                file_path,
                read_only=True,
                data_only=True
            )
            
            all_sheet_names = workbook.sheetnames
            sheets_to_process = sheet_names or all_sheet_names
            
            for sheet_name in sheets_to_process:
                if sheet_name not in all_sheet_names:
                    self._logger.warning(
                        f"Sheet '{sheet_name}' not found in {file_path}"
                    )
                    continue
                
                worksheet = workbook[sheet_name]
                
                # Process sheet in chunks
                for chunk in self._extract_sheet_chunks(
                    worksheet, sheet_name, memory_warnings
                ):
                    total_rows += chunk.row_count
                    chunks_processed += 1
                    yield chunk
            
            workbook.close()
            
        except Exception as e:
            self._logger.error(f"Streaming extraction failed: {e}")
            raise ExtractionError(
                f"Failed to extract file in streaming mode: {e}",
                details={"file_path": file_path, "error": str(e)}
            )
        
        # Return summary
        return StreamingExtractionResult(
            file_path=file_path,
            file_size_mb=file_size_mb,
            sheet_names=all_sheet_names,
            total_rows=total_rows,
            chunks_processed=chunks_processed,
            used_streaming=True,
            memory_warnings=memory_warnings,
            quality=self._calculate_quality(total_rows, chunks_processed)
        )
    
    def _extract_sheet_chunks(
        self,
        worksheet: Worksheet,
        sheet_name: str,
        memory_warnings: list[str]
    ) -> Generator[StreamingChunk, None, None]:
        """
        Extract a worksheet in chunks.
        
        Args:
            worksheet: The worksheet to extract.
            sheet_name: Name of the worksheet.
            memory_warnings: List to append memory warnings to.
        
        Yields:
            StreamingChunk objects for each chunk of data.
        
        Supports Requirement 40.2: Chunked processing for sheets >1M rows.
        """
        chunk_size = self._config.chunk_size_rows
        headers: list[str] = []
        chunk_index = 0
        current_row = 1
        is_first_chunk = True
        
        # Estimate total rows (may not be accurate for read-only mode)
        max_row = worksheet.max_row or 0
        total_chunks = (max_row // chunk_size) + 1 if max_row > 0 else None
        
        self._logger.debug(
            f"Extracting sheet '{sheet_name}' with ~{max_row} rows "
            f"in chunks of {chunk_size}"
        )
        
        while True:
            # Check memory pressure
            if self._config.enable_memory_monitoring:
                memory_status = self._memory_monitor.get_status(
                    self._config.max_memory_mb
                )
                
                if memory_status.is_under_pressure:
                    warning = (
                        f"Memory pressure detected at {memory_status.current_mb:.1f}MB "
                        f"({memory_status.usage_percent:.1f}% of limit)"
                    )
                    self._logger.warning(warning)
                    memory_warnings.append(warning)
                    
                    if self._config.graceful_degradation:
                        # Reduce chunk size to lower memory usage
                        chunk_size = max(100, chunk_size // 2)
                        self._logger.info(
                            f"Reducing chunk size to {chunk_size} due to memory pressure"
                        )
            
            # Extract chunk
            chunk_data: list[list[Any]] = []
            end_row = current_row + chunk_size - 1
            
            for row in worksheet.iter_rows(
                min_row=current_row,
                max_row=end_row,
                values_only=True
            ):
                if row is None:
                    continue
                
                row_values = list(row)
                
                # Detect headers from first row
                if is_first_chunk and not headers and any(row_values):
                    headers = [
                        str(v) if v is not None else f"Column_{i}"
                        for i, v in enumerate(row_values)
                    ]
                    continue
                
                # Skip empty rows
                if any(v is not None for v in row_values):
                    chunk_data.append(row_values)
            
            # Check if we've reached the end
            is_last_chunk = len(chunk_data) < chunk_size or end_row >= max_row
            
            if chunk_data or is_first_chunk:
                yield StreamingChunk(
                    sheet_name=sheet_name,
                    start_row=current_row,
                    end_row=current_row + len(chunk_data) - 1,
                    headers=headers,
                    data=chunk_data,
                    is_first_chunk=is_first_chunk,
                    is_last_chunk=is_last_chunk,
                    chunk_index=chunk_index,
                    total_chunks=total_chunks
                )
            
            if is_last_chunk:
                break
            
            current_row = end_row + 1
            chunk_index += 1
            is_first_chunk = False
    
    def _calculate_quality(
        self,
        total_rows: int,
        chunks_processed: int
    ) -> ExtractionQuality:
        """
        Calculate extraction quality metrics.
        
        Args:
            total_rows: Total rows extracted.
            chunks_processed: Number of chunks processed.
            
        Returns:
            ExtractionQuality with computed metrics.
        """
        has_data = total_rows > 0
        
        return ExtractionQuality(
            score=0.8 if has_data else 0.0,
            data_completeness=1.0 if has_data else 0.0,
            structure_clarity=0.7,  # Streaming mode has less structure info
            has_headers=True,
            has_data=has_data,
            error_count=0,
            warning_count=0
        )


# =============================================================================
# Chunk Aggregator
# =============================================================================

class ChunkAggregator:
    """
    Aggregates streaming chunks into complete sheet data.
    
    Collects chunks from streaming extraction and combines them into
    complete ExtractedSheetData objects when all chunks are received.
    
    Attributes:
        _sheets: Dictionary mapping sheet names to accumulated data.
        _headers: Dictionary mapping sheet names to headers.
    """
    
    def __init__(self) -> None:
        """Initialize the chunk aggregator."""
        self._sheets: dict[str, list[list[Any]]] = {}
        self._headers: dict[str, list[str]] = {}
        self._logger = logging.getLogger(__name__)
    
    def add_chunk(self, chunk: StreamingChunk) -> Optional[ExtractedSheetData]:
        """
        Add a chunk and return complete sheet data if last chunk.
        
        Args:
            chunk: The streaming chunk to add.
            
        Returns:
            ExtractedSheetData if this was the last chunk, None otherwise.
        """
        sheet_name = chunk.sheet_name
        
        # Initialize sheet data if first chunk
        if chunk.is_first_chunk:
            self._sheets[sheet_name] = []
            self._headers[sheet_name] = chunk.headers
        
        # Accumulate data
        if sheet_name in self._sheets:
            self._sheets[sheet_name].extend(chunk.data)
        
        # Return complete data if last chunk
        if chunk.is_last_chunk:
            return self._finalize_sheet(sheet_name)
        
        return None
    
    def _finalize_sheet(self, sheet_name: str) -> ExtractedSheetData:
        """
        Finalize and return complete sheet data.
        
        Args:
            sheet_name: Name of the sheet to finalize.
            
        Returns:
            Complete ExtractedSheetData for the sheet.
        """
        data = self._sheets.pop(sheet_name, [])
        headers = self._headers.pop(sheet_name, [])
        
        return ExtractedSheetData(
            sheet_name=sheet_name,
            headers=headers,
            data=data,
            row_count=len(data),
            column_count=len(headers),
            has_headers=bool(headers)
        )
    
    def get_pending_sheets(self) -> list[str]:
        """
        Get list of sheets with pending (incomplete) data.
        
        Returns:
            List of sheet names with incomplete data.
        """
        return list(self._sheets.keys())
    
    def clear(self) -> None:
        """Clear all accumulated data."""
        self._sheets.clear()
        self._headers.clear()


# =============================================================================
# Factory Function
# =============================================================================

def create_streaming_extractor(
    chunk_size_rows: int = DEFAULT_CHUNK_SIZE_ROWS,
    max_memory_mb: int = DEFAULT_MAX_MEMORY_MB,
    large_file_threshold_mb: int = DEFAULT_LARGE_FILE_THRESHOLD_MB
) -> StreamingExcelExtractor:
    """
    Create a configured streaming extractor.
    
    Factory function for creating StreamingExcelExtractor with common
    configuration options.
    
    Args:
        chunk_size_rows: Number of rows per chunk.
        max_memory_mb: Maximum memory usage in MB.
        large_file_threshold_mb: File size threshold for streaming.
        
    Returns:
        Configured StreamingExcelExtractor instance.
    """
    config = StreamingExtractionConfig(
        chunk_size_rows=chunk_size_rows,
        max_memory_mb=max_memory_mb,
        large_file_threshold_mb=large_file_threshold_mb
    )
    
    monitor = ProcessMemoryMonitor()
    
    return StreamingExcelExtractor(config, monitor)
