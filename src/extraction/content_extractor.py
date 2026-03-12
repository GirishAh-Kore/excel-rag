"""
Content extraction engine for Excel files.

This module provides functionality to parse Excel files (.xlsx, .xls, .xlsm),
extract structured data, handle formulas, formatting, pivot tables, and charts.
"""

import io
import logging
from datetime import datetime
from dateutil import parser as date_parser
from typing import Any, BinaryIO, Dict, List, Optional, Tuple, TYPE_CHECKING, Union

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    xlrd = None

from src.models.domain_models import (
    CellData,
    ChartData,
    DataType,
    PivotTableData,
    SheetData,
    WorkbookData,
)

if TYPE_CHECKING:
    from src.extraction.language_detection import ExcelLanguageDetectionService

logger = logging.getLogger(__name__)


class ExtractionError(Exception):
    """Base exception for extraction errors."""
    pass


class CorruptedFileError(ExtractionError):
    """Exception raised when a file is corrupted."""
    pass


class UnsupportedFormatError(ExtractionError):
    """Exception raised when file format is not supported."""
    pass


class MemoryError(ExtractionError):
    """Exception raised when file is too large to process."""
    pass


class ContentExtractor:
    """
    Extracts structured data from Excel files.
    
    Supports .xlsx files via openpyxl and provides comprehensive extraction
    of cell data, formulas, formatting, pivot tables, and charts.
    """
    
    def __init__(
        self,
        max_rows_per_sheet: int = 10000,
        language_detection_service: Optional["ExcelLanguageDetectionService"] = None
    ):
        """
        Initialize the content extractor.
        
        Args:
            max_rows_per_sheet: Maximum number of rows to process per sheet
            language_detection_service: Optional language detection service.
                Creates default if not provided.
        """
        self.max_rows_per_sheet = max_rows_per_sheet
        self.logger = logging.getLogger(__name__)
        self.failed_files = []  # Track failed files for reporting
        
        # Initialize language detection service
        if language_detection_service is None:
            from src.extraction.language_detection import ExcelLanguageDetectionService
            self._language_service = ExcelLanguageDetectionService()
        else:
            self._language_service = language_detection_service
    
    def extract_workbook(
        self,
        file_content: bytes,
        file_id: str,
        file_name: str,
        file_path: str,
        modified_time: datetime
    ) -> WorkbookData:
        """
        Extract all sheets and metadata from an Excel file.
        
        Supports both .xlsx (via openpyxl) and legacy .xls (via xlrd) formats.
        Implements comprehensive error handling for corrupted files, unsupported
        formats, and memory issues.
        
        Args:
            file_content: Raw bytes of the Excel file
            file_id: Google Drive file ID
            file_name: Name of the file
            file_path: Full path in Google Drive
            modified_time: Last modified timestamp
            
        Returns:
            WorkbookData containing all extracted information (may have partial data)
            
        Raises:
            CorruptedFileError: If file is corrupted
            UnsupportedFormatError: If file format is not supported
            MemoryError: If file is too large to process
        """
        try:
            # Check file size (basic memory protection)
            file_size_mb = len(file_content) / (1024 * 1024)
            if file_size_mb > 100:  # 100 MB limit
                error_msg = f"File {file_name} is too large ({file_size_mb:.2f} MB). Maximum size is 100 MB."
                self.logger.error(error_msg)
                self._track_failed_file(file_id, file_name, "memory_error", error_msg)
                raise MemoryError(error_msg)
            
            # Determine file type and load workbook
            is_xls = file_name.lower().endswith('.xls') and not file_name.lower().endswith('.xlsx')
            
            try:
                if is_xls:
                    # Try to load as legacy .xls file
                    workbook, sheet_names = self._load_xls_workbook(file_content, file_name)
                    is_legacy = True
                else:
                    # Try to load as .xlsx file using openpyxl
                    workbook = self._load_xlsx_workbook(file_content, file_name)
                    sheet_names = workbook.sheetnames
                    is_legacy = False
            except ValueError as e:
                # File loading failed - could be corrupted or unsupported
                error_msg = str(e)
                if "password" in error_msg.lower() or "encrypted" in error_msg.lower():
                    self.logger.error(f"File {file_name} is password-protected")
                    self._track_failed_file(file_id, file_name, "password_protected", error_msg)
                    raise UnsupportedFormatError(f"File {file_name} is password-protected")
                elif "invalid" in error_msg.lower() or "corrupted" in error_msg.lower():
                    self.logger.error(f"File {file_name} is corrupted or invalid")
                    self._track_failed_file(file_id, file_name, "corrupted", error_msg)
                    raise CorruptedFileError(f"File {file_name} is corrupted: {error_msg}")
                else:
                    self.logger.error(f"Unsupported file format for {file_name}")
                    self._track_failed_file(file_id, file_name, "unsupported_format", error_msg)
                    raise UnsupportedFormatError(f"Unsupported file format: {error_msg}")
            
            # Extract all sheets (with error recovery)
            sheets = []
            sheet_errors = []
            
            for sheet_name in sheet_names:
                try:
                    if is_legacy:
                        # For .xls files, we'll need to handle differently in extract_sheet
                        # For now, skip legacy format sheets (will be implemented if needed)
                        self.logger.warning(
                            f"Legacy .xls format not fully supported yet for sheet '{sheet_name}' "
                            f"in {file_name}. Please convert to .xlsx format."
                        )
                        sheet_errors.append(f"Sheet '{sheet_name}': Legacy format not fully supported")
                        continue
                    else:
                        worksheet = workbook[sheet_name]
                        sheet_data = self.extract_sheet(worksheet, file_name)
                        sheets.append(sheet_data)
                        self.logger.info(f"Extracted sheet '{sheet_name}' from {file_name}")
                except MemoryError as e:
                    error_msg = f"Sheet '{sheet_name}' is too large to process"
                    self.logger.error(f"{error_msg}: {e}")
                    sheet_errors.append(error_msg)
                    continue
                except Exception as e:
                    error_msg = f"Failed to extract sheet '{sheet_name}'"
                    self.logger.error(f"{error_msg} from {file_name}: {e}")
                    sheet_errors.append(f"{error_msg}: {str(e)[:100]}")
                    # Continue with other sheets
                    continue
            
            # If no sheets were successfully extracted, raise an error
            if not sheets:
                error_msg = f"No sheets could be extracted from {file_name}"
                if sheet_errors:
                    error_msg += f". Errors: {'; '.join(sheet_errors)}"
                self.logger.error(error_msg)
                self._track_failed_file(file_id, file_name, "no_sheets_extracted", error_msg)
                raise CorruptedFileError(error_msg)
            
            # Detect primary language from extracted sheets
            detected_language = self._detect_workbook_language(sheets)
            
            # Create workbook data (may have partial results)
            workbook_data = WorkbookData(
                file_id=file_id,
                file_name=file_name,
                file_path=file_path,
                sheets=sheets,
                modified_time=modified_time,
                detected_language=detected_language
            )
            
            # Log success with any warnings
            log_msg = (
                f"Successfully extracted workbook {file_name}: "
                f"{len(sheets)} sheets, "
                f"{workbook_data.total_pivot_tables} pivot tables, "
                f"{workbook_data.total_charts} charts, "
                f"language={detected_language}"
            )
            if sheet_errors:
                log_msg += f". Warnings: {len(sheet_errors)} sheets had errors"
            self.logger.info(log_msg)
            
            return workbook_data
            
        except (CorruptedFileError, UnsupportedFormatError, MemoryError):
            # Re-raise our custom exceptions
            raise
        except Exception as e:
            # Catch-all for unexpected errors
            error_msg = f"Unexpected error extracting workbook {file_name}: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            self._track_failed_file(file_id, file_name, "unexpected_error", error_msg)
            raise ExtractionError(error_msg) from e
    
    def _track_failed_file(
        self, file_id: str, file_name: str, error_type: str, error_message: str
    ) -> None:
        """
        Track a failed file for reporting.
        
        Args:
            file_id: File ID
            file_name: File name
            error_type: Type of error
            error_message: Error message
        """
        self.failed_files.append({
            "file_id": file_id,
            "file_name": file_name,
            "error_type": error_type,
            "error_message": error_message,
            "timestamp": datetime.now()
        })
    
    def get_failed_files(self) -> List[Dict[str, Any]]:
        """
        Get list of files that failed extraction.
        
        Returns:
            List of failed file information
        """
        return self.failed_files.copy()
    
    def clear_failed_files(self) -> None:
        """Clear the failed files list."""
        self.failed_files.clear()
    
    def _detect_workbook_language(self, sheets: List[SheetData]) -> str:
        """
        Detect the primary language of workbook content from extracted sheets.
        
        Analyzes headers and data from all sheets to determine the primary
        language of the workbook content.
        
        Args:
            sheets: List of extracted sheet data.
            
        Returns:
            ISO 639-1 language code (e.g., "en", "th", "mixed").
        
        Supports Requirement 23.1: Detect the primary language of Excel content.
        """
        if not sheets:
            return "en"  # Default to English
        
        # Collect text samples from all sheets
        all_headers: List[str] = []
        all_rows: List[Dict[str, Any]] = []
        
        for sheet in sheets:
            if sheet.headers:
                all_headers.extend(sheet.headers)
            if sheet.rows:
                # Sample first 50 rows per sheet
                all_rows.extend(sheet.rows[:50])
        
        # Use language detection service
        detected_language = self._language_service.detect_from_sheet_data(
            headers=all_headers,
            rows=all_rows,
            sample_size=100
        )
        
        self.logger.debug(
            f"Detected language '{detected_language}' from {len(sheets)} sheets"
        )
        
        return detected_language
    
    def _load_xlsx_workbook(self, file_content: bytes, file_name: str) -> openpyxl.Workbook:
        """
        Load an Excel workbook from bytes using openpyxl.
        
        Args:
            file_content: Raw bytes of the Excel file
            file_name: Name of the file (for error messages)
            
        Returns:
            Loaded openpyxl Workbook
            
        Raises:
            ValueError: If file cannot be loaded
        """
        try:
            # Create a BytesIO object from the file content
            file_stream = io.BytesIO(file_content)
            
            # Try to load with openpyxl
            # data_only=True gets calculated values instead of formulas
            # keep_vba=False to avoid loading VBA macros
            workbook = openpyxl.load_workbook(
                file_stream,
                data_only=False,  # We want both formulas and values
                keep_vba=False,
                rich_text=False
            )
            
            return workbook
            
        except openpyxl.utils.exceptions.InvalidFileException as e:
            # Check if it's a password-protected file
            if "password" in str(e).lower() or "encrypted" in str(e).lower():
                raise ValueError(f"File {file_name} is password-protected and cannot be opened")
            raise ValueError(f"Invalid Excel file format for {file_name}: {e}")
        except Exception as e:
            raise ValueError(f"Failed to load Excel file {file_name}: {e}")
    
    def _load_xls_workbook(self, file_content: bytes, file_name: str) -> Tuple[Any, List[str]]:
        """
        Load a legacy .xls Excel workbook from bytes using xlrd.
        
        Args:
            file_content: Raw bytes of the Excel file
            file_name: Name of the file (for error messages)
            
        Returns:
            Tuple of (xlrd.Book, list of sheet names)
            
        Raises:
            ValueError: If file cannot be loaded or xlrd is not available
        """
        if not XLRD_AVAILABLE:
            raise ValueError(
                f"Cannot process legacy .xls file {file_name}: xlrd library not available. "
                "Please convert to .xlsx format or install xlrd."
            )
        
        try:
            # Create a BytesIO object from the file content
            file_stream = io.BytesIO(file_content)
            
            # Load with xlrd
            workbook = xlrd.open_workbook(file_contents=file_content)
            sheet_names = workbook.sheet_names()
            
            return workbook, sheet_names
            
        except xlrd.XLRDError as e:
            # Check if it's a password-protected file
            if "password" in str(e).lower() or "encrypted" in str(e).lower():
                raise ValueError(f"File {file_name} is password-protected and cannot be opened")
            raise ValueError(f"Invalid .xls file format for {file_name}: {e}")
        except Exception as e:
            raise ValueError(f"Failed to load .xls file {file_name}: {e}")
    
    def extract_sheet(self, worksheet: Worksheet, file_name: str) -> SheetData:
        """
        Extract structured data from a single sheet.
        
        Detects headers, extracts cell values with data type preservation,
        handles merged cells, and limits processing to max_rows_per_sheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            file_name: Name of the file (for logging)
            
        Returns:
            SheetData containing extracted information
        """
        sheet_name = worksheet.title
        
        # Get sheet dimensions
        max_row = min(worksheet.max_row or 0, self.max_rows_per_sheet)
        max_col = worksheet.max_column or 0
        
        if max_row == 0 or max_col == 0:
            # Empty sheet
            return SheetData(
                sheet_name=sheet_name,
                headers=[],
                rows=[],
                data_types={},
                row_count=0,
                column_count=0,
                summary=f"Empty sheet '{sheet_name}' from {file_name}",
                has_dates=False,
                has_numbers=False
            )
        
        # Detect header row (analyze first 5 rows)
        header_row_idx = self._detect_header_row(worksheet, max_col)
        
        # Extract headers
        headers = self._extract_headers(worksheet, header_row_idx, max_col)
        
        # Extract data rows with formula handling
        rows = []
        data_type_samples = {header: [] for header in headers}
        
        for row_idx in range(header_row_idx + 1, max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                # Extract cell data including formula information
                cell_data = self._extract_cell_data(cell, worksheet)
                
                # For the row data, use the calculated value if available, otherwise the raw value
                if cell_data.is_formula and cell_data.value is not None:
                    # Use calculated value for formulas
                    row_data[header] = cell_data.value
                else:
                    row_data[header] = cell_data.value
                
                # Collect data type samples (use the underlying data type, not FORMULA)
                if cell_data.data_type != DataType.EMPTY:
                    # For formulas, use the type of the calculated value
                    if cell_data.is_formula and cell_data.value is not None:
                        # Infer type from calculated value
                        if isinstance(cell_data.value, bool):
                            data_type_samples[header].append(DataType.BOOLEAN)
                        elif isinstance(cell_data.value, (int, float)):
                            data_type_samples[header].append(DataType.NUMBER)
                        elif isinstance(cell_data.value, datetime):
                            data_type_samples[header].append(DataType.DATE)
                        else:
                            data_type_samples[header].append(DataType.TEXT)
                    else:
                        data_type_samples[header].append(cell_data.data_type)
            
            # Only add non-empty rows
            if any(v is not None and v != "" for v in row_data.values()):
                rows.append(row_data)
        
        # Infer data types for each column
        data_types = self._infer_data_types(data_type_samples)
        
        # Detect if sheet has dates or numbers
        has_dates = DataType.DATE in data_types.values()
        has_numbers = DataType.NUMBER in data_types.values()
        
        # Generate summary
        summary = self._generate_sheet_summary(
            sheet_name, file_name, headers, len(rows), has_dates, has_numbers
        )
        
        # Extract pivot tables and charts (will be implemented in later tasks)
        pivot_tables = self.extract_pivot_tables(worksheet)
        charts = self.extract_charts(worksheet)
        
        return SheetData(
            sheet_name=sheet_name,
            headers=headers,
            rows=rows,
            data_types=data_types,
            row_count=len(rows),
            column_count=len(headers),
            summary=summary,
            has_dates=has_dates,
            has_numbers=has_numbers,
            pivot_tables=pivot_tables,
            charts=charts
        )
    
    def _detect_header_row(self, worksheet: Worksheet, max_col: int) -> int:
        """
        Detect the header row by analyzing the first 5 rows for text-heavy content.
        
        Args:
            worksheet: openpyxl Worksheet object
            max_col: Maximum column index
            
        Returns:
            Row index (1-based) of the detected header row
        """
        # Analyze first 5 rows
        max_rows_to_check = min(5, worksheet.max_row or 0)
        
        best_row = 1
        best_score = 0
        
        for row_idx in range(1, max_rows_to_check + 1):
            text_count = 0
            non_empty_count = 0
            
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if cell.value is not None and str(cell.value).strip():
                    non_empty_count += 1
                    # Check if it's text (not a number or date)
                    if isinstance(cell.value, str):
                        text_count += 1
            
            # Score: prefer rows with more text cells and more non-empty cells
            score = text_count * 2 + non_empty_count
            
            if score > best_score:
                best_score = score
                best_row = row_idx
        
        return best_row
    
    def _extract_headers(self, worksheet: Worksheet, header_row: int, max_col: int) -> List[str]:
        """
        Extract column headers from the header row.
        
        Args:
            worksheet: openpyxl Worksheet object
            header_row: Row index of headers (1-based)
            max_col: Maximum column index
            
        Returns:
            List of header names
        """
        headers = []
        
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            
            # Handle merged cells - get the value from the merged range
            if cell.coordinate in worksheet.merged_cells:
                # Find the merged range
                for merged_range in worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Get the top-left cell of the merged range
                        top_left_cell = worksheet.cell(
                            row=merged_range.min_row,
                            column=merged_range.min_col
                        )
                        header_value = top_left_cell.value
                        break
                else:
                    header_value = cell.value
            else:
                header_value = cell.value
            
            # Convert to string and clean up
            if header_value is None or str(header_value).strip() == "":
                header_name = f"Column_{get_column_letter(col_idx)}"
            else:
                header_name = str(header_value).strip()
            
            headers.append(header_name)
        
        return headers
    
    def _extract_cell_data(self, cell: Cell, worksheet: Worksheet) -> CellData:
        """
        Extract comprehensive cell data including formulas, values, and errors.
        
        Args:
            cell: openpyxl Cell object
            worksheet: openpyxl Worksheet object
            
        Returns:
            CellData object with all cell information
        """
        # Handle merged cells
        actual_cell = cell
        if cell.coordinate in worksheet.merged_cells:
            for merged_range in worksheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the top-left cell of the merged range
                    actual_cell = worksheet.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col
                    )
                    break
        
        # Check if cell has a formula
        formula_text = None
        formula_error = None
        is_formula = False
        
        # openpyxl stores formulas in the data_type attribute
        if hasattr(actual_cell, 'data_type') and actual_cell.data_type == 'f':
            is_formula = True
            # Get the formula text
            if hasattr(actual_cell, 'value') and isinstance(actual_cell.value, str):
                formula_text = actual_cell.value
        
        # Get the cell value
        cell_value = actual_cell.value
        
        # Check for formula errors
        if isinstance(cell_value, str) and cell_value.startswith('#'):
            # Common Excel errors
            error_types = ['#DIV/0!', '#REF!', '#VALUE!', '#N/A', '#NAME?', '#NUM!', '#NULL!']
            for error_type in error_types:
                if cell_value.startswith(error_type):
                    formula_error = error_type
                    break
        
        # Determine data type
        if cell_value is None or cell_value == "":
            data_type = DataType.EMPTY
        elif is_formula:
            data_type = DataType.FORMULA
        elif isinstance(cell_value, bool):
            data_type = DataType.BOOLEAN
        elif isinstance(cell_value, (int, float)):
            data_type = DataType.NUMBER
        elif isinstance(cell_value, datetime):
            data_type = DataType.DATE
        else:
            data_type = DataType.TEXT
        
        # Get cell format
        cell_format = None
        if hasattr(actual_cell, 'number_format'):
            cell_format = actual_cell.number_format
        
        # Normalize date values to ISO 8601 if it's a date
        if data_type == DataType.DATE and isinstance(cell_value, datetime):
            # Keep the datetime object but ensure it's properly formatted
            # The ISO 8601 conversion will happen when serializing
            pass
        
        return CellData(
            value=cell_value,
            data_type=data_type,
            formula=formula_text,
            formula_error=formula_error,
            format=cell_format,
            is_formula=is_formula
        )
    
    def _extract_cell_value(
        self, cell: Cell, worksheet: Worksheet
    ) -> Tuple[Any, DataType]:
        """
        Extract value and data type from a cell (simplified version).
        
        This is a simpler version used for header extraction.
        
        Args:
            cell: openpyxl Cell object
            worksheet: openpyxl Worksheet object
            
        Returns:
            Tuple of (cell_value, data_type)
        """
        cell_data = self._extract_cell_data(cell, worksheet)
        return cell_data.value, cell_data.data_type
    
    def _infer_data_types(self, data_type_samples: Dict[str, List[DataType]]) -> Dict[str, DataType]:
        """
        Infer the primary data type for each column based on samples.
        
        Args:
            data_type_samples: Dictionary mapping column names to lists of observed data types
            
        Returns:
            Dictionary mapping column names to inferred data types
        """
        data_types = {}
        
        for column, samples in data_type_samples.items():
            if not samples:
                data_types[column] = DataType.TEXT
                continue
            
            # Count occurrences of each type
            type_counts = {}
            for dtype in samples:
                type_counts[dtype] = type_counts.get(dtype, 0) + 1
            
            # Choose the most common type
            most_common_type = max(type_counts, key=type_counts.get)
            data_types[column] = most_common_type
        
        return data_types
    
    def _generate_sheet_summary(
        self,
        sheet_name: str,
        file_name: str,
        headers: List[str],
        row_count: int,
        has_dates: bool,
        has_numbers: bool
    ) -> str:
        """
        Generate a natural language summary of the sheet.
        
        Args:
            sheet_name: Name of the sheet
            file_name: Name of the file
            headers: List of column headers
            row_count: Number of data rows
            has_dates: Whether sheet contains dates
            has_numbers: Whether sheet contains numbers
            
        Returns:
            Natural language summary string
        """
        summary_parts = [f"Sheet '{sheet_name}' from {file_name}"]
        
        if headers:
            summary_parts.append(f"with {len(headers)} columns: {', '.join(headers[:5])}")
            if len(headers) > 5:
                summary_parts[-1] += f" and {len(headers) - 5} more"
        
        summary_parts.append(f"containing {row_count} rows of data")
        
        data_types = []
        if has_numbers:
            data_types.append("numerical values")
        if has_dates:
            data_types.append("dates")
        
        if data_types:
            summary_parts.append(f"including {' and '.join(data_types)}")
        
        return " ".join(summary_parts) + "."
    
    def extract_pivot_tables(self, worksheet: Worksheet) -> List[PivotTableData]:
        """
        Extract pivot table definitions and data from a sheet.
        
        Note: openpyxl has limited support for pivot tables. We can access
        pivot table definitions but not the calculated results directly.
        
        Args:
            worksheet: openpyxl Worksheet object
            
        Returns:
            List of PivotTableData objects
        """
        pivot_tables = []
        
        try:
            # Access pivot tables via worksheet._pivots
            if not hasattr(worksheet, '_pivots') or not worksheet._pivots:
                return []
            
            for idx, pivot in enumerate(worksheet._pivots):
                try:
                    pivot_data = self._extract_pivot_table(pivot, idx, worksheet)
                    if pivot_data:
                        pivot_tables.append(pivot_data)
                except Exception as e:
                    self.logger.warning(f"Failed to extract pivot table {idx} from sheet {worksheet.title}: {e}")
                    continue
        
        except Exception as e:
            self.logger.warning(f"Failed to access pivot tables in sheet {worksheet.title}: {e}")
        
        return pivot_tables
    
    def _extract_pivot_table(
        self, pivot: Any, index: int, worksheet: Worksheet
    ) -> Optional[PivotTableData]:
        """
        Extract a single pivot table's information.
        
        Args:
            pivot: openpyxl pivot table object
            index: Index of the pivot table
            worksheet: openpyxl Worksheet object
            
        Returns:
            PivotTableData object or None if extraction fails
        """
        try:
            # Get pivot table name
            name = getattr(pivot, 'name', f"PivotTable{index + 1}")
            
            # Get location (where the pivot table is displayed)
            location = "Unknown"
            if hasattr(pivot, 'location'):
                loc = pivot.location
                if hasattr(loc, 'ref'):
                    location = loc.ref
            
            # Get source range
            source_range = "Unknown"
            if hasattr(pivot, 'cache') and hasattr(pivot.cache, 'cacheSource'):
                cache_source = pivot.cache.cacheSource
                if hasattr(cache_source, 'worksheetSource'):
                    ws_source = cache_source.worksheetSource
                    if hasattr(ws_source, 'ref'):
                        source_range = ws_source.ref
            
            # Extract field information
            row_fields = []
            column_fields = []
            data_fields = []
            filters = {}
            
            # Row fields
            if hasattr(pivot, 'rowFields') and pivot.rowFields:
                for field in pivot.rowFields:
                    if hasattr(field, 'x'):
                        # x is the field index
                        field_name = self._get_pivot_field_name(pivot, field.x)
                        if field_name:
                            row_fields.append(field_name)
            
            # Column fields
            if hasattr(pivot, 'colFields') and pivot.colFields:
                for field in pivot.colFields:
                    if hasattr(field, 'x'):
                        field_name = self._get_pivot_field_name(pivot, field.x)
                        if field_name:
                            column_fields.append(field_name)
            
            # Data fields
            if hasattr(pivot, 'dataFields') and pivot.dataFields:
                for data_field in pivot.dataFields:
                    if hasattr(data_field, 'name'):
                        data_fields.append(data_field.name)
                    elif hasattr(data_field, 'fld'):
                        field_name = self._get_pivot_field_name(pivot, data_field.fld)
                        if field_name:
                            # Include aggregation type if available
                            agg_type = getattr(data_field, 'subtotal', 'Sum')
                            data_fields.append(f"{agg_type} of {field_name}")
            
            # Generate summary
            summary = self._generate_pivot_summary(
                name, row_fields, column_fields, data_fields
            )
            
            return PivotTableData(
                name=name,
                location=location,
                source_range=source_range,
                row_fields=row_fields,
                column_fields=column_fields,
                data_fields=data_fields,
                filters=filters,
                aggregated_data={},  # We can't easily get calculated results
                summary=summary
            )
        
        except Exception as e:
            self.logger.error(f"Error extracting pivot table {index}: {e}")
            return None
    
    def _get_pivot_field_name(self, pivot: Any, field_index: int) -> Optional[str]:
        """
        Get the name of a pivot field by its index.
        
        Args:
            pivot: openpyxl pivot table object
            field_index: Index of the field
            
        Returns:
            Field name or None
        """
        try:
            if hasattr(pivot, 'cache') and hasattr(pivot.cache, 'cacheFields'):
                cache_fields = pivot.cache.cacheFields
                if field_index < len(cache_fields):
                    field = cache_fields[field_index]
                    if hasattr(field, 'name'):
                        return field.name
        except Exception:
            pass
        
        return None
    
    def _generate_pivot_summary(
        self,
        name: str,
        row_fields: List[str],
        column_fields: List[str],
        data_fields: List[str]
    ) -> str:
        """
        Generate a natural language summary of a pivot table.
        
        Args:
            name: Pivot table name
            row_fields: List of row grouping fields
            column_fields: List of column grouping fields
            data_fields: List of aggregated data fields
            
        Returns:
            Natural language summary
        """
        parts = [f"Pivot table '{name}'"]
        
        if data_fields:
            parts.append(f"showing {', '.join(data_fields)}")
        
        grouping_parts = []
        if row_fields:
            grouping_parts.append(f"grouped by {', '.join(row_fields)}")
        if column_fields:
            grouping_parts.append(f"across {', '.join(column_fields)}")
        
        if grouping_parts:
            parts.append(" ".join(grouping_parts))
        
        return " ".join(parts) + "."
    
    def extract_charts(self, worksheet: Worksheet) -> List[ChartData]:
        """
        Extract chart metadata and source data from a sheet.
        
        Args:
            worksheet: openpyxl Worksheet object
            
        Returns:
            List of ChartData objects
        """
        charts = []
        
        try:
            # Access charts via worksheet._charts
            if not hasattr(worksheet, '_charts') or not worksheet._charts:
                return []
            
            for idx, chart in enumerate(worksheet._charts):
                try:
                    chart_data = self._extract_chart(chart, idx, worksheet)
                    if chart_data:
                        charts.append(chart_data)
                except Exception as e:
                    self.logger.warning(f"Failed to extract chart {idx} from sheet {worksheet.title}: {e}")
                    continue
        
        except Exception as e:
            self.logger.warning(f"Failed to access charts in sheet {worksheet.title}: {e}")
        
        return charts
    
    def _extract_chart(
        self, chart: Any, index: int, worksheet: Worksheet
    ) -> Optional[ChartData]:
        """
        Extract a single chart's information.
        
        Args:
            chart: openpyxl chart object
            index: Index of the chart
            worksheet: openpyxl Worksheet object
            
        Returns:
            ChartData object or None if extraction fails
        """
        try:
            # Get chart name
            name = getattr(chart, 'name', f"Chart{index + 1}")
            
            # Get chart type
            chart_type = chart.__class__.__name__.replace('Chart', '').lower()
            # Common types: bar, line, pie, scatter, area, etc.
            
            # Get chart title
            title = None
            if hasattr(chart, 'title') and chart.title:
                title = str(chart.title)
            
            # Get axis labels
            x_axis_label = None
            y_axis_label = None
            
            if hasattr(chart, 'x_axis') and chart.x_axis:
                if hasattr(chart.x_axis, 'title') and chart.x_axis.title:
                    x_axis_label = str(chart.x_axis.title)
            
            if hasattr(chart, 'y_axis') and chart.y_axis:
                if hasattr(chart.y_axis, 'title') and chart.y_axis.title:
                    y_axis_label = str(chart.y_axis.title)
            
            # Extract series information
            series_list = []
            source_ranges = []
            
            if hasattr(chart, 'series'):
                for series in chart.series:
                    series_info = {}
                    
                    # Series title/name
                    if hasattr(series, 'title') and series.title:
                        series_info['name'] = str(series.title)
                    
                    # Series values reference
                    if hasattr(series, 'val') and series.val:
                        val_ref = str(series.val)
                        series_info['values'] = val_ref
                        source_ranges.append(val_ref)
                    
                    # Series categories/labels reference
                    if hasattr(series, 'cat') and series.cat:
                        cat_ref = str(series.cat)
                        series_info['categories'] = cat_ref
                        if cat_ref not in source_ranges:
                            source_ranges.append(cat_ref)
                    
                    if series_info:
                        series_list.append(series_info)
            
            # Combine source ranges
            source_range = ", ".join(source_ranges) if source_ranges else "Unknown"
            
            # Generate summary
            summary = self._generate_chart_summary(
                chart_type, title, x_axis_label, y_axis_label, len(series_list)
            )
            
            return ChartData(
                name=name,
                chart_type=chart_type,
                title=title,
                source_range=source_range,
                series=series_list,
                x_axis_label=x_axis_label,
                y_axis_label=y_axis_label,
                summary=summary
            )
        
        except Exception as e:
            self.logger.error(f"Error extracting chart {index}: {e}")
            return None
    
    def _generate_chart_summary(
        self,
        chart_type: str,
        title: Optional[str],
        x_axis_label: Optional[str],
        y_axis_label: Optional[str],
        series_count: int
    ) -> str:
        """
        Generate a natural language summary of a chart.
        
        Args:
            chart_type: Type of chart (bar, line, pie, etc.)
            title: Chart title
            x_axis_label: X-axis label
            y_axis_label: Y-axis label
            series_count: Number of data series
            
        Returns:
            Natural language summary
        """
        parts = [f"{chart_type.capitalize()} chart"]
        
        if title:
            parts.append(f"titled '{title}'")
        
        if series_count > 0:
            parts.append(f"with {series_count} data series")
        
        axis_parts = []
        if x_axis_label:
            axis_parts.append(f"X-axis: {x_axis_label}")
        if y_axis_label:
            axis_parts.append(f"Y-axis: {y_axis_label}")
        
        if axis_parts:
            parts.append(f"({', '.join(axis_parts)})")
        
        return " ".join(parts) + "."
    
    def format_cell_value(self, cell_data: CellData) -> str:
        """
        Format a cell value according to its Excel format.
        
        Args:
            cell_data: CellData object with value and format
            
        Returns:
            Formatted string representation
        """
        if cell_data.value is None:
            return ""
        
        # Handle formula errors
        if cell_data.formula_error:
            return cell_data.formula_error
        
        # Handle dates
        if cell_data.data_type == DataType.DATE and isinstance(cell_data.value, datetime):
            return cell_data.value.isoformat()
        
        # Handle numbers with formatting
        if cell_data.data_type == DataType.NUMBER and cell_data.format:
            format_str = cell_data.format.lower()
            
            # Currency format
            if '$' in format_str or '€' in format_str or '£' in format_str:
                currency_symbol = '$'
                if '€' in format_str:
                    currency_symbol = '€'
                elif '£' in format_str:
                    currency_symbol = '£'
                return f"{currency_symbol}{cell_data.value:,.2f}"
            
            # Percentage format
            elif '%' in format_str:
                return f"{cell_data.value * 100:.2f}%"
            
            # Thousands separator
            elif ',' in format_str:
                return f"{cell_data.value:,.2f}"
        
        # Default: convert to string
        return str(cell_data.value)
    
    def parse_date_from_format(self, value: Any, cell_format: Optional[str]) -> Optional[datetime]:
        """
        Parse a date value and normalize to datetime object.
        
        Args:
            value: Cell value (could be string, number, or datetime)
            cell_format: Excel number format string
            
        Returns:
            datetime object or None if not a date
        """
        if isinstance(value, datetime):
            return value
        
        # Check if format indicates a date
        if cell_format:
            date_indicators = ['d', 'm', 'y', 'h', 's']
            format_lower = cell_format.lower()
            if any(indicator in format_lower for indicator in date_indicators):
                # Try to parse as date
                try:
                    if isinstance(value, str):
                        return date_parser.parse(value)
                    elif isinstance(value, (int, float)):
                        # Excel date serial number
                        from datetime import timedelta
                        # Excel epoch is 1899-12-30
                        excel_epoch = datetime(1899, 12, 30)
                        return excel_epoch + timedelta(days=value)
                except (ValueError, TypeError):
                    pass
        
        return None
    
    def generate_embeddings_text(self, sheet_data: SheetData, file_name: str) -> List[str]:
        """
        Generate text chunks for embedding from sheet data.
        
        Creates multiple text representations of the sheet for better semantic search:
        1. File + sheet name + headers
        2. Sheet summary with sample data
        3. Column-wise summaries for numerical data
        4. Pivot table descriptions
        5. Chart descriptions
        
        Args:
            sheet_data: Extracted sheet data
            file_name: Name of the file
            
        Returns:
            List of text chunks suitable for embedding
        """
        chunks = []
        
        # Chunk 1: Basic metadata with headers
        metadata_chunk = self._generate_metadata_chunk(sheet_data, file_name)
        chunks.append(metadata_chunk)
        
        # Chunk 2: Sheet summary with sample data
        if sheet_data.rows:
            summary_chunk = self._generate_summary_chunk(sheet_data, file_name)
            chunks.append(summary_chunk)
        
        # Chunk 3: Column-wise summaries for numerical columns
        column_chunks = self._generate_column_chunks(sheet_data, file_name)
        chunks.extend(column_chunks)
        
        # Chunk 4: Pivot table descriptions
        for pivot in sheet_data.pivot_tables:
            pivot_chunk = self._generate_pivot_chunk(pivot, sheet_data.sheet_name, file_name)
            chunks.append(pivot_chunk)
        
        # Chunk 5: Chart descriptions
        for chart in sheet_data.charts:
            chart_chunk = self._generate_chart_chunk(chart, sheet_data.sheet_name, file_name)
            chunks.append(chart_chunk)
        
        return chunks
    
    def _generate_metadata_chunk(self, sheet_data: SheetData, file_name: str) -> str:
        """Generate metadata chunk with file, sheet, and header information."""
        parts = [
            f"File: {file_name}",
            f"Sheet: {sheet_data.sheet_name}",
        ]
        
        if sheet_data.headers:
            parts.append(f"Columns: {', '.join(sheet_data.headers)}")
        
        parts.append(f"Contains {sheet_data.row_count} rows of data")
        
        # Add data type information
        data_info = []
        if sheet_data.has_numbers:
            data_info.append("numerical data")
        if sheet_data.has_dates:
            data_info.append("dates")
        if sheet_data.has_pivot_tables:
            data_info.append(f"{len(sheet_data.pivot_tables)} pivot tables")
        if sheet_data.has_charts:
            data_info.append(f"{len(sheet_data.charts)} charts")
        
        if data_info:
            parts.append(f"Includes: {', '.join(data_info)}")
        
        return ". ".join(parts) + "."
    
    def _generate_summary_chunk(self, sheet_data: SheetData, file_name: str) -> str:
        """Generate summary chunk with sample data from first 5 rows."""
        parts = [
            f"Data from {file_name}, sheet '{sheet_data.sheet_name}':",
        ]
        
        # Include first 5 rows as sample
        sample_rows = sheet_data.rows[:5]
        
        for idx, row in enumerate(sample_rows, start=1):
            row_parts = []
            for header, value in row.items():
                if value is not None and value != "":
                    # Format the value appropriately
                    if isinstance(value, datetime):
                        value_str = value.strftime("%Y-%m-%d")
                    elif isinstance(value, float):
                        value_str = f"{value:.2f}"
                    else:
                        value_str = str(value)
                    row_parts.append(f"{header}: {value_str}")
            
            if row_parts:
                parts.append(f"Row {idx}: {', '.join(row_parts)}")
        
        return " ".join(parts) + "."
    
    def _generate_column_chunks(self, sheet_data: SheetData, file_name: str) -> List[str]:
        """Generate chunks for numerical columns with statistics."""
        chunks = []
        
        # Focus on numerical columns
        numerical_columns = [
            col for col, dtype in sheet_data.data_types.items()
            if dtype == DataType.NUMBER
        ]
        
        for column in numerical_columns:
            # Calculate basic statistics
            values = [
                row.get(column) for row in sheet_data.rows
                if row.get(column) is not None and isinstance(row.get(column), (int, float))
            ]
            
            if not values:
                continue
            
            total = sum(values)
            avg = total / len(values)
            min_val = min(values)
            max_val = max(values)
            
            chunk = (
                f"Column '{column}' in {file_name}, sheet '{sheet_data.sheet_name}': "
                f"Total = {total:.2f}, Average = {avg:.2f}, "
                f"Min = {min_val:.2f}, Max = {max_val:.2f}, "
                f"Based on {len(values)} values."
            )
            chunks.append(chunk)
        
        return chunks
    
    def _generate_pivot_chunk(
        self, pivot: PivotTableData, sheet_name: str, file_name: str
    ) -> str:
        """Generate embedding chunk for a pivot table."""
        parts = [
            f"Pivot table in {file_name}, sheet '{sheet_name}':",
            pivot.summary
        ]
        
        if pivot.row_fields:
            parts.append(f"Row grouping: {', '.join(pivot.row_fields)}")
        
        if pivot.column_fields:
            parts.append(f"Column grouping: {', '.join(pivot.column_fields)}")
        
        if pivot.data_fields:
            parts.append(f"Aggregations: {', '.join(pivot.data_fields)}")
        
        return " ".join(parts) + "."
    
    def _generate_chart_chunk(
        self, chart: ChartData, sheet_name: str, file_name: str
    ) -> str:
        """Generate embedding chunk for a chart."""
        parts = [
            f"Chart in {file_name}, sheet '{sheet_name}':",
            chart.summary
        ]
        
        if chart.title:
            parts.append(f"Title: {chart.title}")
        
        if chart.series:
            series_names = [s.get('name', 'Unnamed') for s in chart.series]
            parts.append(f"Data series: {', '.join(series_names)}")
        
        return " ".join(parts) + "."
