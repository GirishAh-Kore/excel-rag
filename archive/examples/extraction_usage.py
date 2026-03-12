"""
Example usage of the Content Extraction module.

This script demonstrates how to extract data from Excel files.
"""

import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.extraction import (
    ContentExtractor,
    CorruptedFileError,
    UnsupportedFormatError,
)


def create_sample_excel() -> bytes:
    """Create a sample Excel file for demonstration."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    
    # Add headers
    ws['A1'] = "Product"
    ws['B1'] = "Quantity"
    ws['C1'] = "Price"
    ws['D1'] = "Total"
    ws['E1'] = "Date"
    
    # Add data
    ws['A2'] = "Laptop"
    ws['B2'] = 5
    ws['C2'] = 999.99
    ws['D2'] = '=B2*C2'  # Formula
    ws['E2'] = datetime(2024, 1, 15)
    
    ws['A3'] = "Mouse"
    ws['B3'] = 20
    ws['C3'] = 25.50
    ws['D3'] = '=B3*C3'  # Formula
    ws['E3'] = datetime(2024, 1, 16)
    
    ws['A4'] = "Keyboard"
    ws['B4'] = 10
    ws['C4'] = 75.00
    ws['D4'] = '=B4*C4'  # Formula
    ws['E4'] = datetime(2024, 1, 17)
    
    # Format currency column
    ws['C2'].number_format = '$#,##0.00'
    ws['C3'].number_format = '$#,##0.00'
    ws['C4'].number_format = '$#,##0.00'
    ws['D2'].number_format = '$#,##0.00'
    ws['D3'].number_format = '$#,##0.00'
    ws['D4'].number_format = '$#,##0.00'
    
    # Save to bytes
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream.read()


def main():
    """Demonstrate content extraction."""
    print("=" * 60)
    print("Content Extraction Example")
    print("=" * 60)
    
    # Create sample Excel file
    print("\n1. Creating sample Excel file...")
    file_content = create_sample_excel()
    print(f"   Created file with {len(file_content)} bytes")
    
    # Initialize extractor
    print("\n2. Initializing ContentExtractor...")
    extractor = ContentExtractor(max_rows_per_sheet=10000)
    
    # Extract workbook
    print("\n3. Extracting workbook...")
    try:
        workbook_data = extractor.extract_workbook(
            file_content=file_content,
            file_id="sample_123",
            file_name="sales_data.xlsx",
            file_path="/examples/sales_data.xlsx",
            modified_time=datetime.now()
        )
        
        print(f"   ✓ Successfully extracted workbook")
        print(f"   - File: {workbook_data.file_name}")
        print(f"   - Sheets: {len(workbook_data.sheets)}")
        print(f"   - Pivot tables: {workbook_data.total_pivot_tables}")
        print(f"   - Charts: {workbook_data.total_charts}")
        
    except (CorruptedFileError, UnsupportedFormatError) as e:
        print(f"   ✗ Extraction failed: {e}")
        return
    
    # Display sheet information
    print("\n4. Sheet Details:")
    for sheet in workbook_data.sheets:
        print(f"\n   Sheet: {sheet.sheet_name}")
        print(f"   - Headers: {', '.join(sheet.headers)}")
        print(f"   - Rows: {sheet.row_count}")
        print(f"   - Columns: {sheet.column_count}")
        print(f"   - Has numbers: {sheet.has_numbers}")
        print(f"   - Has dates: {sheet.has_dates}")
        
        # Display sample data
        print(f"\n   Sample Data (first 3 rows):")
        for idx, row in enumerate(sheet.rows[:3], start=1):
            print(f"   Row {idx}:")
            for header, value in row.items():
                if value is not None:
                    print(f"      {header}: {value}")
    
    # Generate embedding text
    print("\n5. Generating Embedding Text:")
    for sheet in workbook_data.sheets:
        chunks = extractor.generate_embeddings_text(sheet, workbook_data.file_name)
        print(f"\n   Generated {len(chunks)} text chunks for '{sheet.sheet_name}':")
        for idx, chunk in enumerate(chunks, start=1):
            print(f"\n   Chunk {idx}:")
            print(f"   {chunk[:200]}..." if len(chunk) > 200 else f"   {chunk}")
    
    # Check for failed files
    print("\n6. Failed Files:")
    failed_files = extractor.get_failed_files()
    if failed_files:
        for failed in failed_files:
            print(f"   - {failed['file_name']}: {failed['error_type']}")
    else:
        print("   No failed files")
    
    print("\n" + "=" * 60)
    print("Extraction Complete!")
    print("=" * 60)


if __name__ == "__main__":
    main()
